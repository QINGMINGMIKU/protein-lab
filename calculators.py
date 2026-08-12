"""
蛋白质计算核心：MW、消光系数（Biopython ProtParam）、浓度、BLI 稀释规划
"""
from dataclasses import dataclass
from typing import List

from Bio.SeqUtils.ProtParam import ProteinAnalysis


def sanitize_seq(sequence: str) -> str:
    """清洗序列：去换行/空格/终止符/非氨基酸字符"""
    import re
    seq = sequence.upper()
    seq = re.sub(r'\s+', '', seq)
    seq = seq.replace('*', '')
    seq = re.sub(r'[^ACDEFGHIKLMNPQRSTVWY]', '', seq)
    return seq


def calc_mw(sequence: str) -> float:
    """蛋白质分子量 (Da) — 使用 Biopython ProtParam"""
    return ProteinAnalysis(sanitize_seq(sequence)).molecular_weight()


def calc_ext_coeff(sequence: str) -> dict:
    """计算还原态和氧化态消光系数 — 使用 Biopython ProtParam (Pace et al. 1995)"""
    seq = sanitize_seq(sequence)
    pa = ProteinAnalysis(seq)
    ext_red, ext_ox = pa.molar_extinction_coefficient()  # → (reduced, oxidized)
    mw = pa.molecular_weight()
    nW = seq.count("W")
    nY = seq.count("Y")
    nC = seq.count("C")
    abs_0_1pct = ext_ox / mw if mw > 0 else 0
    return {
        "mw": round(mw, 1),
        "nW": nW, "nY": nY, "nC": nC,
        "ext_red": round(ext_red, 0),
        "ext_ox": round(ext_ox, 0),
        "abs_0_1pct": round(abs_0_1pct, 4),
    }


def calc_conc(a280: float, ext_coeff: float, mw: float,
              path_length: float = 1.0) -> dict:
    """Beer-Lambert: A = ε·c·l → c"""
    if ext_coeff <= 0:
        raise ValueError("消光系数为 0，无法用 A280 定量（序列不含 W/Y/C）")
    if path_length <= 0:
        raise ValueError(f"光程必须 > 0，当前值: {path_length} cm")
    if a280 < 0:
        raise ValueError(f"A280 不能为负数，当前值: {a280}")
    molar_M = a280 / (ext_coeff * path_length)
    molar_conc_uM = molar_M * 1e6
    mass_conc_ng_uL = molar_conc_uM * mw / 1000  # 1 µM × MW/1000 = ng/µL
    return {
        "a280": a280,
        "path_length_cm": path_length,
        "epsilon": ext_coeff,
        "mw": mw,
        "molar_conc_uM": round(molar_conc_uM, 2),
        "molar_conc_nM": round(molar_conc_uM * 1e3, 2),
        "molar_conc_M": round(molar_conc_uM / 1e6, 10),
        "mass_conc_mg_mL": round(mass_conc_ng_uL / 1000, 4),
        "mass_conc_ug_mL": round(mass_conc_ng_uL, 2),
        "mass_conc_ng_uL": round(mass_conc_ng_uL, 2),
    }


# ═══════════════════════════════════════════════════════════
#  浓度单位换算 kernel（隐藏能力：6 单位互转）
# ═══════════════════════════════════════════════════════════
# canonical 基准：molar→µM，mass→ng/µL。跨 kind（摩尔↔质量）必须提供 mw (Da)：
#   molar→mass: base × mw / 1000；mass→molar: base × 1000 / mw
# 前端 static/app.js 有逐行镜像 convertConc()，改动时两边同步。

CONC_UNITS = {
    "M":     {"kind": "molar", "factor": 1e6},    # → µM
    "uM":    {"kind": "molar", "factor": 1},
    "nM":    {"kind": "molar", "factor": 1e-3},
    "mg/mL": {"kind": "mass",  "factor": 1000},   # → ng/µL
    "ug/mL": {"kind": "mass",  "factor": 1},
    "ng/uL": {"kind": "mass",  "factor": 1},
}


def convert_concentration(value: float, from_unit: str, to_unit: str,
                          mw: float = None) -> float:
    """6 种浓度单位互转。同 kind 直接比例换算；跨 kind（摩尔↔质量）需 mw (Da)。"""
    if from_unit not in CONC_UNITS:
        raise ValueError(f"未知单位: {from_unit}")
    if to_unit not in CONC_UNITS:
        raise ValueError(f"未知单位: {to_unit}")
    f, t = CONC_UNITS[from_unit], CONC_UNITS[to_unit]
    base = value * f["factor"]  # canonical 基准（µM 或 ng/µL）
    if f["kind"] != t["kind"]:
        if not mw or mw <= 0:
            raise ValueError("跨摩尔/质量换算需要分子量 mw (Da)")
        base = base * mw / 1000 if f["kind"] == "molar" else base * 1000 / mw
    return base / t["factor"]


@dataclass
class DilutionStep:
    step: int
    conc_uM: float
    stock_vol_uL: float
    buffer_vol_uL: float
    total_vol_uL: float


def calc_dilution_series(stock_conc_uM: float, start_conc_uM: float,
                         dilution_factor: float, n_steps: int,
                         vol_per_well_uL: float,
                         extra_dead_vol_uL: float = 0.0) -> List[DilutionStep]:
    """
    BLI 梯度稀释规划 — 连续递推稀释

    每个步骤需要足够体积来：(a) 取 vol_per_well_uL 到孔中 +
    (b) 留足下一步稀释所需的母液 + (c) 死体积裕量。

    从最后一步向前递推：第 i 步总体积 = 孔体积 + 第 i+1 步总体积/稀释倍数 + 死体积。
    递推后每步总体积向上取整到 5 μL 倍数，保证移液量尽量整洁。

    Parameters
    ----------
    stock_conc_uM : 母液浓度 (μM)
    start_conc_uM : 起始最高浓度 (μM), 必须 ≤ stock_conc_uM
    dilution_factor : 稀释倍数 (如 2 表示 2 倍稀释)
    n_steps : 梯度步数
    vol_per_well_uL : 每孔所需体积 (μL)
    extra_dead_vol_uL : 额外死体积 (μL), 用于保证移液准确

    Returns
    -------
    list of DilutionStep
    """
    if stock_conc_uM <= 0:
        raise ValueError(f"母液浓度必须 > 0，当前值: {stock_conc_uM} μM")
    if start_conc_uM <= 0:
        raise ValueError(f"起始浓度必须 > 0，当前值: {start_conc_uM} μM")
    if start_conc_uM > stock_conc_uM:
        raise ValueError(
            f"起始浓度 ({start_conc_uM} μM) 不能超过母液浓度 ({stock_conc_uM} μM)")
    if dilution_factor <= 1:
        raise ValueError(f"稀释倍数必须 > 1，当前值: {dilution_factor}")
    if n_steps < 1:
        raise ValueError(f"梯度步数必须 ≥ 1，当前值: {n_steps}")
    if vol_per_well_uL <= 0:
        raise ValueError(f"每孔体积必须 > 0，当前值: {vol_per_well_uL} μL")

    # 从后向前递推，计算第 0 步所需最大体积
    import math
    total_last = vol_per_well_uL + extra_dead_vol_uL
    total_cur = total_last
    for _ in range(n_steps - 1):
        total_cur = vol_per_well_uL + total_cur / dilution_factor + extra_dead_vol_uL
    # 向上取整到 100 μL，所有步骤统一体积
    uniform_total = math.ceil(total_cur / 100) * 100

    steps = []
    for i in range(n_steps):
        target_conc = start_conc_uM / (dilution_factor ** i)
        total_needed = uniform_total

        if i == 0:
            # 第一步: 直接从母液配制
            stock_vol = total_needed * target_conc / stock_conc_uM
            buffer_vol = total_needed - stock_vol
        else:
            # 后续步骤: 从上一步的剩余液中取 total_needed / factor
            stock_vol = total_needed / dilution_factor
            buffer_vol = total_needed - stock_vol

        steps.append(DilutionStep(
            step=i + 1,
            conc_uM=round(target_conc, 4),
            stock_vol_uL=round(stock_vol, 2),
            buffer_vol_uL=round(buffer_vol, 2),
            total_vol_uL=round(total_needed, 2),
        ))
    return steps


# ═══════════════════════════════════════════════════════════
#  酶活动力学计算
# ═══════════════════════════════════════════════════════════

import numpy as np
import openpyxl

ROW_ORDER = "ABCDEFGH"


def parse_tecan_xlsx(filepath: str) -> dict:
    """解析 TECAN Spark xlsx，返回 {meta: {...}, wells: {A1: {times:[], od:[]}, ...}}"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    meta = {"sample": "", "wavelength": "", "temps": []}
    wells = {}

    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v is None:
            continue
        label = str(v).strip().rstrip(":")

        if label == "Name":
            meta["sample"] = str(ws.cell(r, 2).value or "")
        elif label == "Measurement wavelength":
            meta["wavelength"] = ws.cell(r, 5).value
        elif label == "Target temperature":
            meta["target_temp"] = ws.cell(r, 5).value

        elif label == "Cycle Nr.":
            time_s = float(ws.cell(r - 2, 2).value)
            meta.setdefault("temps", []).append(time_s)
            header_row = r + 1
            col_map = {}
            for c in range(2, 14):
                vh = ws.cell(header_row, c).value
                if vh is not None:
                    col_map[c] = int(vh)
            for rr in range(header_row + 1, header_row + 9):
                rl = str(ws.cell(rr, 1).value or "").strip()
                if rl not in ROW_ORDER:
                    continue
                for c, wcol in col_map.items():
                    vv = ws.cell(rr, c).value
                    if vv is None or str(vv).strip() == "":
                        continue
                    key = f"{rl}{wcol}"
                    wells.setdefault(key, {"times": [], "od": []})
                    wells[key]["times"].append(time_s)
                    wells[key]["od"].append(float(vv))
    wb.close()
    return {"meta": meta, "wells": {k: v for k, v in sorted(wells.items())}}


def fit_kinetics(times: list, od: list) -> dict:
    """线性拟合 → ΔOD/min、R²"""
    t = np.asarray(times, float)
    od_arr = np.asarray(od, float)
    n = len(t)
    if n < 2:
        return {"slope": None, "intercept": None, "r2": None, "n": n}
    k, b = np.polyfit(t, od_arr, 1)
    dod_min = round(float(k * 60), 6)
    if n > 2:
        pred = k * t + b
        ss_res = float(np.sum((od_arr - pred) ** 2))
        ss_tot = float(np.sum((od_arr - od_arr.mean()) ** 2))
        r2 = round(float(1 - ss_res / ss_tot), 4) if ss_tot > 0 else None
    else:
        r2 = None
    return {"slope": dod_min, "intercept": round(float(b), 6), "r2": r2, "n": n}
