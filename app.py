"""
Protein Lab — 本地蛋白质实验管理系统
Flask 主应用
"""
import sys
import os
import json
import hashlib
import tempfile
from datetime import datetime
import webbrowser
from io import BytesIO
from threading import Timer, Event, Lock
from flask import Flask, request, jsonify, render_template, send_file
from openpyxl import Workbook
from openpyxl.styles import Font

# 确保能找到同目录模块
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import numpy as np
import paths
import models
import services
from calculators import calc_ext_coeff, calc_conc, calc_dilution_series, sanitize_seq, parse_tecan_xlsx, fit_kinetics, sub_blank, align_wells, snap_ylim, correct_slopes, aggregate_groups

app = Flask(__name__,
            template_folder=paths.resource_path("templates"),
            static_folder=paths.resource_path("static"))


# ── 静态资源版本号（缓存破坏）───────────────────────────────
# 取 app.js / style.css 的修改时间戳做 ?v= 参数：改前端代码后浏览器自动拉新文件，
# 不再被旧缓存坑（新增 tab 的按钮调新函数，旧 JS 缓存里没有 → 点击无效）。
@app.context_processor
def inject_static_version():
    import time as _time
    try:
        app_js = os.path.getmtime(os.path.join(app.static_folder, "app.js"))
        style_css = os.path.getmtime(os.path.join(app.static_folder, "style.css"))
        version = str(int(max(app_js, style_css)))
    except OSError:
        version = "0"
    return {"static_version": version}

# ── Undo 栈（内存中，最多 20 条）──────────────────────────
_undo_stack = []


def _push_undo(item_type: str, data: dict):
    _undo_stack.append({"type": item_type, "data": data})
    if len(_undo_stack) > 20:
        _undo_stack.pop(0)


def _pop_undo():
    return _undo_stack.pop() if _undo_stack else None


# ═══════════════════════════════════════════════════════════
#  页面路由
# ═══════════════════════════════════════════════════════════

@app.route("/")
def index():
    return render_template("proteins.html")


@app.route("/proteins")
def page_proteins():
    return render_template("proteins.html")


@app.route("/calculator")
def page_calculator():
    return render_template("calculator.html")


@app.route("/experiments")
def page_experiments():
    return render_template("experiments.html", exp_types=models.EXP_TYPES)


@app.route("/experiments/<int:eid>")
def page_experiment_detail(eid):
    e = models.exp_get(eid)
    if not e:
        return "实验不存在", 404
    # models.exp_get 已反序列化 params/results（见 models._json_unwrap），这里只兜底成 dict 供模板
    for field in ("params", "results"):
        val = e.get(field)
        e[field] = val if isinstance(val, dict) else {}
    raws = models.exp_raw_list(eid, with_version=True)
    return render_template("experiment_detail.html", exp=e, exp_types=models.EXP_TYPES,
                           raws=raws)


# ═══════════════════════════════════════════════════════════
#  Proteins API
# ═══════════════════════════════════════════════════════════

@app.route("/api/proteins", methods=["GET"])
def api_protein_list():
    search = request.args.get("q", "")
    tag_filter = request.args.get("tag", "")
    proteins = models.protein_list(search, tag_filter)
    return jsonify(proteins)


@app.route("/api/proteins/tags", methods=["GET"])
def api_protein_tags():
    return jsonify(models.protein_tags())


@app.route("/api/proteins/<int:pid>", methods=["GET"])
def api_protein_get(pid):
    p = models.protein_get(pid)
    if not p:
        return jsonify({"error": "蛋白不存在"}), 404
    return jsonify(p)


@app.route("/api/proteins", methods=["POST"])
def api_protein_create():
    data = request.get_json()
    name = data.get("name", "").strip()
    sequence = sanitize_seq(data.get("sequence", ""))
    if not name or not sequence:
        return jsonify({"error": "名称和序列不能为空"}), 400

    # 检查重复
    existing = models.protein_get_by_name(name)
    if existing:
        return jsonify({"error": f"蛋白 '{name}' 已存在"}), 409

    # 计算 MW / ε
    c = calc_ext_coeff(sequence)
    pid = models.protein_create(
        name=name, sequence=sequence,
        tag=data.get("tag", ""), notes=data.get("notes", ""),
        mw=c["mw"], nW=c["nW"], nY=c["nY"], nC=c["nC"],
        ext_red=c["ext_red"], ext_ox=c["ext_ox"], abs_0_1pct=c["abs_0_1pct"],
    )
    return jsonify(models.protein_get(pid)), 201


@app.route("/api/proteins/<int:pid>", methods=["PUT"])
def api_protein_update(pid):
    data = request.get_json()
    kwargs = {}
    for field in ["name", "sequence", "tag", "notes"]:
        if field in data:
            val = data[field]
            if isinstance(val, str):
                val = val.strip()
                if field == "sequence":
                    val = val.upper()
            kwargs[field] = val
    # 如果更新了序列，重新计算
    if "sequence" in kwargs:
        c = calc_ext_coeff(kwargs["sequence"])
        kwargs.update(mw=c["mw"], nW=c["nW"], nY=c["nY"], nC=c["nC"],
                      ext_red=c["ext_red"], ext_ox=c["ext_ox"],
                      abs_0_1pct=c["abs_0_1pct"])
    models.protein_update(pid, **kwargs)
    return jsonify(models.protein_get(pid))


@app.route("/api/proteins/<int:pid>", methods=["DELETE"])
def api_protein_delete(pid):
    p = models.protein_get(pid)
    if p:
        _push_undo("protein", dict(p))
    models.protein_delete(pid)
    return jsonify({"ok": True})


@app.route("/api/proteins/batch-delete", methods=["POST"])
def api_protein_batch_delete():
    ids = request.get_json().get("ids", [])
    deleted = 0
    for pid in ids:
        p = models.protein_get(int(pid))
        if p:
            _push_undo("protein", dict(p))
            models.protein_delete(int(pid))
            deleted += 1
    return jsonify({"ok": True, "deleted": deleted})


@app.route("/api/proteins/batch-tags", methods=["POST"])
def api_protein_batch_tags():
    """批量修改标签：为多个蛋白添加/移除标签（不覆盖已有标签）"""
    data = request.get_json()
    ids = data.get("ids", [])
    add_set = {t.strip() for t in data.get("add", "").split(",") if t.strip()}
    remove_set = {t.strip() for t in data.get("remove", "").split(",") if t.strip()}
    if not add_set and not remove_set:
        return jsonify({"error": "请至少添加或移除一个标签"}), 400
    updated = 0
    for pid in ids:
        p = models.protein_get(int(pid))
        if not p:
            continue
        cur = {t.strip() for t in (p.get("tag") or "").split(",") if t.strip()}
        cur.difference_update(remove_set)
        cur.update(add_set)
        models.protein_update(int(pid), tag=", ".join(sorted(cur)))
        updated += 1
    return jsonify({"ok": True, "updated": updated})


@app.route("/api/proteins/delete-all", methods=["POST"])
def api_protein_delete_all():
    proteins = models.protein_list()
    for p in proteins:
        _push_undo("protein", dict(p))
    models.protein_delete_all()
    return jsonify({"ok": True, "deleted": len(proteins)})


@app.route("/api/proteins/refresh-all", methods=["POST"])
def api_protein_refresh_all():
    """一键重算所有蛋白的 MW 和消光系数"""
    proteins = models.protein_list()
    count = 0
    for p in proteins:
        c = calc_ext_coeff(p["sequence"])
        models.protein_update(p["id"],
            mw=c["mw"], nW=c["nW"], nY=c["nY"], nC=c["nC"],
            ext_red=c["ext_red"], ext_ox=c["ext_ox"], abs_0_1pct=c["abs_0_1pct"])
        count += 1
    return jsonify({"ok": True, "refreshed": count})


@app.route("/api/proteins/import", methods=["POST"])
def api_protein_import_fasta():
    """批量导入 FASTA 文本"""
    data = request.get_json()
    fasta_text = data.get("fasta", "")
    tag = data.get("tag", "")
    notes = data.get("notes", "")
    if not fasta_text.strip():
        return jsonify({"error": "FASTA 内容为空"}), 400

    imported = []
    skipped = []
    entries = parse_fasta(fasta_text)
    for header, seq in entries:
        if not seq:
            continue
        name = header.strip()
        existing = models.protein_get_by_name(name)
        if existing:
            skipped.append(name)
            continue
        c = calc_ext_coeff(seq)
        models.protein_create(
            name=name, sequence=seq, tag=tag, notes=notes,
            mw=c["mw"], nW=c["nW"], nY=c["nY"], nC=c["nC"],
            ext_red=c["ext_red"], ext_ox=c["ext_ox"], abs_0_1pct=c["abs_0_1pct"],
        )
        imported.append(name)

    return jsonify({"imported": imported, "skipped": skipped, "total": len(entries)})


def parse_fasta(text: str) -> list:
    """解析 FASTA 文本，返回 [(header, sequence), ...]"""
    entries = []
    current_header = ""
    current_seq = []
    for line in text.strip().splitlines():
        line = line.strip()
        if not line:
            continue
        if line.startswith(">"):
            if current_header:
                entries.append((current_header, "".join(current_seq)))
            current_header = line[1:]  # 去掉 >
            current_seq = []
        else:
            current_seq.append(line)
    if current_header:
        entries.append((current_header, "".join(current_seq)))
    return entries


# ═══════════════════════════════════════════════════════════
#  Calculation API
# ═══════════════════════════════════════════════════════════

@app.route("/api/calc/concentration", methods=["POST"])
def api_calc_concentration():
    data = request.get_json()
    protein_id = data.get("protein_id")
    a280 = data.get("a280")
    oxidized = data.get("oxidized", True)
    path_length = data.get("path_length_cm", 1.0)

    if not protein_id or a280 is None:
        # 允许直接传序列
        sequence = data.get("sequence")
        if not sequence:
            return jsonify({"error": "需要 protein_id 或 sequence"}), 400
        c = calc_ext_coeff(sequence)
        mw = c["mw"]
        epsilon = c["ext_ox"] if oxidized else c["ext_red"]
    else:
        p = models.protein_get(int(protein_id))
        if not p:
            return jsonify({"error": "蛋白不存在"}), 404
        epsilon = p["ext_ox"] if oxidized else p["ext_red"]
        mw = p["mw"]

    try:
        path_length = float(path_length)
        result = calc_conc(float(a280), epsilon, mw, path_length)
    except (ValueError, TypeError) as e:
        return jsonify({"error": str(e)}), 400
    return jsonify(result)


@app.route("/api/calc/dilution", methods=["POST"])
def api_calc_dilution():
    data = request.get_json()
    try:
        stock_conc = float(data["stock_conc_uM"])
        start_conc = float(data["start_conc_uM"])
        dilution_factor = float(data.get("dilution_factor", 2))
        n_steps = int(data.get("n_steps", 8))
        vol_per_well = float(data["vol_per_well_uL"])
        dead_vol = float(data.get("dead_vol_uL", 5))
    except (KeyError, ValueError) as e:
        return jsonify({"error": f"参数无效: {e}"}), 400

    try:
        steps = calc_dilution_series(stock_conc, start_conc, dilution_factor,
                                     n_steps, vol_per_well, dead_vol)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    return jsonify({
        "stock_conc_uM": stock_conc,
        "dilution_factor": dilution_factor,
        "n_steps": n_steps,
        "vol_per_well_uL": vol_per_well,
        "steps": [{
            "step": s.step,
            "conc_uM": s.conc_uM,
            "stock_vol_uL": s.stock_vol_uL,
            "buffer_vol_uL": s.buffer_vol_uL,
            "total_vol_uL": s.total_vol_uL,
        } for s in steps]
    })


# ═══════════════════════════════════════════════════════════
#  Experiments API
# ═══════════════════════════════════════════════════════════

@app.route("/api/experiments", methods=["GET"])
def api_exp_list():
    exp_type = request.args.get("type", "")
    try:
        limit = int(request.args.get("limit", 50))
    except (ValueError, TypeError):
        limit = 50
    return jsonify(models.exp_list(exp_type, limit))


@app.route("/api/experiments/<int:eid>", methods=["GET"])
def api_exp_get(eid):
    e = models.exp_get(eid)
    if not e:
        return jsonify({"error": "实验不存在"}), 404
    # 附原始数据快照 id（供「从实验复制」等按需拉取 payload，避免大字段常驻详情响应）
    e["_raw_ids"] = [r["id"] for r in models.exp_raw_list(e["id"])]
    return jsonify(e)


@app.route("/api/experiments/<int:eid>/raw/<int:rid>", methods=["GET"])
def api_exp_raw_get(eid, rid):
    """单条原始数据快照（含 payload）——只读，供复制重建会话等场景按需拉取。"""
    raw = models.exp_raw_get(rid)
    if not raw or raw.get("experiment_id") != eid:
        return jsonify({"error": "原始快照不存在"}), 404
    return jsonify(_json_safe(raw))


@app.route("/api/experiments/next-name", methods=["GET"])
def api_exp_next_name():
    """获取自动命名（供前端 prompt 默认值 / 输入框占位）"""
    exp_type = request.args.get("exp_type", "").strip()
    if not exp_type:
        return jsonify({"error": "缺少 exp_type"}), 400
    return jsonify({"name": services.auto_exp_name(exp_type, request.args.get("date", ""))})


@app.route("/api/experiments", methods=["POST"])
def api_exp_create():
    data = request.get_json()
    try:
        e = services.create_experiment(
            title=data.get("title", ""),
            exp_type=data.get("exp_type", ""),
            protein_ids=data.get("protein_ids", []),
            date=data.get("date", ""),
            params=data.get("params", {}),
            results=data.get("results", {}),
            notes=data.get("notes", ""),
        )
    except ValueError as err:
        return jsonify({"error": str(err)}), 400
    return jsonify(e), 201


@app.route("/api/experiments/from-calculation", methods=["POST"])
def api_exp_from_calc():
    """从计算工具一键保存为实验"""
    data = request.get_json()
    # 将计算参数和结果打包进 params/results
    calc_params = data.get("calc_params", {})
    calc_result = data.get("calc_result", {})
    params = {"calc_type": data.get("calc_type", ""), **calc_params}
    try:
        e = services.create_experiment(
            title=data.get("title", ""),
            exp_type=data.get("exp_type", ""),
            protein_ids=data.get("protein_ids", []),
            date=data.get("date", ""),
            params=params, results=calc_result,
            notes=data.get("notes", ""),
        )
    except ValueError as err:
        return jsonify({"error": str(err)}), 400
    return jsonify(e), 201


@app.route("/api/experiments/<int:eid>", methods=["PUT"])
def api_exp_update(eid):
    """编辑实验元数据（标题、类型、日期、备注）"""
    data = request.get_json()
    kwargs = {}
    for key in ["title", "exp_type", "date", "notes", "params", "results"]:
        if key in data:
            kwargs[key] = data[key]
    if kwargs:
        models.exp_update(eid, **kwargs)
    return jsonify(models.exp_get(eid))


@app.route("/api/experiments/<int:eid>/proteins", methods=["PUT"])
def api_exp_update_proteins(eid):
    """单独更新实验的蛋白关联"""
    data = request.get_json()
    protein_ids = data.get("protein_ids", [])
    if isinstance(protein_ids, list):
        protein_ids = services.coerce_int_list(protein_ids)
    models.exp_update(eid, protein_ids=protein_ids)
    return jsonify(models.exp_get(eid))


def _exp_undo_payload(e: dict) -> dict:
    """实验删除进 undo 栈前，附带其原始数据快照 id 列表——恢复时重挂孤儿 raw（FK 已 SET NULL）。"""
    d = dict(e)
    d["_raw_ids"] = [r["id"] for r in models.exp_raw_list(e["id"])]
    return d


@app.route("/api/experiments/<int:eid>", methods=["DELETE"])
def api_exp_delete(eid):
    e = models.exp_get(eid)
    if e:
        _push_undo("experiment", _exp_undo_payload(e))
    models.exp_delete(eid)
    return jsonify({"ok": True})


@app.route("/api/experiments/batch-delete", methods=["POST"])
def api_exp_batch_delete():
    ids = request.get_json().get("ids", [])
    deleted = 0
    for eid in ids:
        e = models.exp_get(int(eid))
        if e:
            _push_undo("experiment", _exp_undo_payload(e))
            models.exp_delete(int(eid))
            deleted += 1
    return jsonify({"ok": True, "deleted": deleted})


@app.route("/api/experiments/delete-all", methods=["POST"])
def api_exp_delete_all():
    exps = models.exp_list(limit=9999)
    for e in exps:
        _push_undo("experiment", _exp_undo_payload(e))
    models.exp_delete_all()
    return jsonify({"ok": True, "deleted": len(exps)})


# ═══════════════════════════════════════════════════════════
#  Undo API
# ═══════════════════════════════════════════════════════════

@app.route("/api/undo/status", methods=["GET"])
def api_undo_status():
    return jsonify({"count": len(_undo_stack)})


@app.route("/api/undo", methods=["POST"])
def api_undo_restore():
    item = _pop_undo()
    if not item:
        return jsonify({"error": "无撤销项"}), 404
    data = item["data"]
    if item["type"] == "protein":
        existing = models.protein_get_by_name(data["name"])
        if existing:
            return jsonify({"error": f"蛋白 '{data['name']}' 已存在，无法撤销"}), 409
        models.protein_create(
            name=data["name"], sequence=data["sequence"],
            tag=data.get("tag", ""), notes=data.get("notes", ""),
            mw=data.get("mw", 0), nW=data.get("nW", 0),
            nY=data.get("nY", 0), nC=data.get("nC", 0),
            ext_red=data.get("ext_red", 0), ext_ox=data.get("ext_ox", 0),
            abs_0_1pct=data.get("abs_0_1pct", 0),
        )
        return jsonify({"ok": True, "restored": data["name"]})
    elif item["type"] == "experiment":
        protein_ids = data.get("protein_ids", [])
        new_eid = models.exp_create(
            title=data["title"], exp_type=data["exp_type"],
            protein_ids=protein_ids,
            date=data.get("date", ""),
            params=data.get("params", {}),
            results=data.get("results", {}),
            notes=data.get("notes", ""),
        )
        # 删除时 FK 把 raw.experiment_id 置 NULL；恢复后重挂回新 id，详情页快照不丢
        raw_ids = data.get("_raw_ids") or []
        if raw_ids:
            models.exp_raw_relink(raw_ids, new_eid)
        return jsonify({"ok": True, "restored": data["title"]})
    return jsonify({"error": "未知类型"}), 400


@app.route("/api/experiments/<int:eid>/export", methods=["GET"])
def api_exp_export_single(eid):
    """导出单个实验为 Excel"""
    e = models.exp_get(eid)
    if not e:
        return jsonify({"error": "实验不存在"}), 404
    safe_name = e["title"].replace("/", "_").replace("\\", "_")[:60]
    return _export_excel([e], download_name=f"{safe_name}.xlsx")


@app.route("/api/experiments/export", methods=["GET"])
def api_exp_export():
    """导出全部（或按类型筛选）实验为 Excel"""
    exp_type = request.args.get("type", "")
    exps = models.exp_list(exp_type, limit=9999)
    return _export_excel(exps)


def _enzyme_wide_columns(wells, prefix="") -> tuple[list, list]:
    """把每孔时间序列转成宽格式：每孔独立两列（时间 + OD），跨孔并列。
    返回 (headers, col_data)：headers 每孔一对 `{标签} 时间 (min)` / `{标签} OD`；
    标签优先样品名、缺省用孔位，撞名时退回孔位，prefix 用于多实验合并区分来源。
    col_data 为 [(times, ods), ...]，写入时每孔占两列、行按各自时间索引对齐（不足留空）。"""
    headers = []
    col_data = []
    if not isinstance(wells, dict):
        return headers, col_data
    seen = set()
    for wid, w in sorted(wells.items()):
        if not isinstance(w, dict):
            continue
        times = w.get("times") or []
        ods = w.get("od") or []
        if not times or not ods:
            continue
        lbl = (w.get("name") or "").strip() or wid
        full = f"{prefix}{lbl}"
        if full in seen:
            full = f"{prefix}{wid}"
        seen.add(full)
        headers += [f"{full} 时间 (min)", f"{full} OD"]
        col_data.append((times, ods))
    return headers, col_data


def _write_wide_ws(ws, headers, col_data):
    """把宽格式列对写入 worksheet：第 1 行表头加粗 + 固定列宽，数据每孔两列
    （时间 (min) + OD），各列对从第 2 行起按各自时间索引对齐（点数不同不足处留空）。"""
    from openpyxl.utils import get_column_letter
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)
        ws.column_dimensions[get_column_letter(c)].width = 15
    for ci, (times, ods) in enumerate(col_data):
        base = ci * 2 + 1
        for r, (t, od) in enumerate(zip(times, ods), 2):
            ws.cell(row=r, column=base).value = round(t / 60, 3)
            ws.cell(row=r, column=base + 1).value = round(od, 4) if od is not None else ""


def _write_wide_ws_pairs(ws, headers, col_data, fmt=None):
    """通用宽格式列对写入：第 1 行表头加粗 + 固定列宽 15，每单位两列从第 2 行起
    按各自索引对齐写入（点数不同不足处留空）。col_data 为 [(xs, ys), ...]；
    fmt: (x, y) -> (v1, v2) 数值转换，缺省原样 float。
    BLI/AKTA 作图数据用此写入器；酶活专用变换（时间 /60、OD 取 4 位）见 _write_wide_ws。"""
    from openpyxl.utils import get_column_letter
    if fmt is None:
        fmt = lambda x, y: (float(x), float(y))
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)
        ws.column_dimensions[get_column_letter(c)].width = 15
    for ci, (xs, ys) in enumerate(col_data):
        base = ci * 2 + 1
        for r, (x, y) in enumerate(zip(xs, ys), 2):
            v1, v2 = fmt(x, y)
            ws.cell(row=r, column=base).value = v1
            ws.cell(row=r, column=base + 1).value = v2


def _export_excel(exps, download_name="实验记录.xlsx"):
    """共享导出逻辑 — exps 是已查询的实验列表。
    每个实验一个区块：实验名称/日期/类型 各占一行（不再作为表头列），下面接该类型的数据表头 + 数据行。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "实验记录"

    # 根据实验类型选择数据表头（实验名称/日期/类型 单独成行，不进表头）
    def _get_calc_type(e):
        p = e.get("params") or {}
        return p.get("calc_type", "") if isinstance(p, dict) else ""
    all_enzyme = exps and all(_get_calc_type(e) == "enzyme" for e in exps)
    enzyme_wide_groups = []  # (headers, col_data) 宽格式（每孔两列），全酶活导出时附加为作图 Sheet

    CONC_HEADERS = ["蛋白", "MW (Da)", "ε", "Abs 0.1%", "A280", "浓度 (μM)", "浓度 (mg/mL)",
                    "目标浓度 (μM)", "目标体积 (μL)", "取母液 (μL)", "加缓冲液 (μL)"]
    DIL_HEADERS = ["蛋白", "母液 (μM)", "稀释倍数", "每孔体积 (μL)", "步骤",
                   "浓度 (μM)", "总体积 (μL)", "取上步 (μL)", "加缓冲液 (μL)"]
    ENZ_HEADERS = ["孔位/命名", "参考类型", "浓度 (ng/mL)", "浓度 (μM)",
                   "ΔOD/min", "R²", "样本", "波长"]

    def _write_block(headers, rows):
        """写一块数据：表头行（加粗）+ 数据行。"""
        if not rows:
            return
        ws.append(headers)
        for c in range(1, len(headers) + 1):
            ws.cell(row=ws.max_row, column=c).font = Font(bold=True)
        for r in rows:
            ws.append(r)

    first_exp = True
    for e in exps:
        params = e.get("params") or {}
        results = e.get("results") or {}
        calc_type = params.get("calc_type", "")
        proteins = params.get("proteins", [])

        # 实验信息：每项独立一行（块与块之间空一行分隔）
        if not first_exp:
            ws.append([])
        first_exp = False
        meta = [("实验名称", e.get("title", "")), ("日期", e.get("date", "")),
                ("类型", e.get("exp_type", ""))]
        if e.get("notes"):
            meta.append(("备注", e["notes"]))
        for label, val in meta:
            ws.append([label, val])
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True)

        if calc_type == "concentration" and proteins:
            rows = []
            for i, prot in enumerate(proteins):
                if not isinstance(prot, dict):
                    continue
                mw_val = prot.get("mw", "")
                eps_val = prot.get("epsilon", "")
                abs_val = prot.get("abs_0_1pct", "")
                # 新格式：全在 params.proteins 里；旧格式：浓度在 results 数组里
                conc_uM = prot.get("conc_uM", "")
                conc_mg = prot.get("conc_mg_mL", "")
                if (not conc_uM) and isinstance(results, list) and i < len(results):
                    r = results[i]
                    if isinstance(r, dict):
                        conc_uM = r.get("conc_uM", conc_uM)
                        conc_mg = r.get("conc_mg_mL", conc_mg)
                rows.append([
                    prot.get("name", ""),
                    f"{mw_val:.1f}" if isinstance(mw_val, (int, float)) and mw_val else mw_val,
                    f"{eps_val:.0f}" if isinstance(eps_val, (int, float)) and eps_val else eps_val,
                    f"{abs_val:.4f}" if isinstance(abs_val, (int, float)) and abs_val else abs_val,
                    prot.get("a280", ""), conc_uM, conc_mg,
                    prot.get("target_conc", ""), prot.get("target_vol", ""),
                    prot.get("take_vol", ""), prot.get("buffer_vol", ""),
                ])
            _write_block(CONC_HEADERS, rows)

        elif calc_type == "dilution":
            # BLI 稀释实验：每蛋白每步一行
            dil_proteins = params.get("proteins", [])
            rows = []
            for prot in dil_proteins:
                if not isinstance(prot, dict):
                    continue
                pid = str(prot.get("id", ""))
                prot_result = results.get(pid, {}) if isinstance(results, dict) else {}
                prot_steps = prot_result.get("steps", []) if isinstance(prot_result, dict) else []
                if not prot_steps:
                    # 旧格式兼容：results.steps 直接是列表
                    prot_steps = results.get("steps", results) if isinstance(results, dict) else (results if isinstance(results, list) else [])
                for st in prot_steps:
                    if not isinstance(st, dict):
                        continue
                    rows.append([
                        prot.get("name", ""), prot.get("stock_uM", ""), prot.get("factor", ""),
                        prot.get("vol", ""), st.get("step", ""), st.get("conc_uM", ""),
                        st.get("total_vol_uL", ""), st.get("stock_vol_uL", ""),
                        st.get("buffer_vol_uL", ""),
                    ])
            if rows:
                _write_block(DIL_HEADERS, rows)
            else:
                ws.append(["关联蛋白", e.get("protein_names", "")])

        elif calc_type == "enzyme":
            # 酶活实验：每孔一行
            wells = params.get("wells") or params.get("well_info") or {}
            emeta = params.get("meta", {})
            if wells:
                rows = []
                for wid, w in sorted(wells.items()):
                    if not isinstance(w, dict):
                        continue
                    fit = w.get("fit") or {}
                    ref_label = {"blank": "空白", "neg": "阴性", "pos": "阳性"}.get(w.get("ref", ""), "")
                    rows.append([
                        f"{wid} {w.get('name', '')}", ref_label,
                        w.get("conc_ng_ml", ""), w.get("conc_uM", ""),
                        fit.get("slope", ""), fit.get("r2", ""),
                        emeta.get("sample", ""), emeta.get("wavelength", ""),
                    ])
                _write_block(ENZ_HEADERS, rows)
                # 汇总 Sheet 的原始数据收进宽格式（每孔时间+OD 两列），多实验用标题前缀区分
                h2, cd2 = _enzyme_wide_columns(wells, prefix=f"{e['title']} ")
                if h2:
                    enzyme_wide_groups.append((h2, cd2))
            else:
                ws.append(["关联蛋白", e.get("protein_names", "")])

        elif proteins and not calc_type:
            # 旧格式：proteins 存在但没标记 calc_type，走旧逻辑
            rows = []
            for i, prot in enumerate(proteins):
                if not isinstance(prot, dict):
                    continue
                name = prot.get("name", "")
                a280 = prot.get("a280", "")
                conc_uM = ""
                conc_mg = ""
                if isinstance(results, list) and i < len(results):
                    r = results[i]
                    conc_uM = r.get("conc_uM", "")
                    conc_mg = r.get("conc_mg_mL", "")
                rows.append([name, "", "", "", a280, conc_uM, conc_mg, "", "", "", ""])
            _write_block(CONC_HEADERS, rows)

        else:
            ws.append(["关联蛋白", e.get("protein_names", "")])

    # 自适应列宽：按内容最大宽度设置（CJK 按 2 个宽度计，上限 40）
    for col in range(1, ws.max_column + 1):
        max_len = 0
        for cell in ws[col]:
            v = cell.value
            if v is None:
                continue
            n = sum(2 if ord(ch) > 127 else 1 for ch in str(v))
            max_len = max(max_len, n)
        ws.column_dimensions[cell.column_letter].width = min(max_len + 2, 40)

    # 酶活导出：附加「作图数据」宽格式 Sheet（每孔独立时间+OD 两列，方便 Origin/Prism 等外部作图）
    if all_enzyme and enzyme_wide_groups:
        ws2 = wb.create_sheet("作图数据")
        headers2, all_cols = [], []
        for h2, cd2 in enzyme_wide_groups:
            headers2 += h2
            all_cols += cd2
        _write_wide_ws(ws2, headers2, all_cols)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=download_name)


# ═══════════════════════════════════════════════════════════
#  启动
# ═══════════════════════════════════════════════════════════

# ═══════════════════════════════════════════════════════════
#  Weblogo API
# ═══════════════════════════════════════════════════════════

# weblogo 结果缓存：同一请求（相同序列+选项）并发共享一次渲染；切页回来命中缓存秒回
_weblogo_lock = Lock()
_weblogo_cache = {}     # hash -> {"image": ..., "positions": ..., "range": ...}
_weblogo_inflight = {}  # hash -> Event


def _render_weblogo(sequences, color_scheme, n_pos, offset):
    """用 logomaker 渲染信息 logo，返回 base64 PNG data URL + 位置信息"""
    import base64
    import pandas as pd
    import logomaker
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    # 中文字体（图内块标题），内部已 use('Agg') + 注册字体
    from fonts import setup_matplotlib_cjk
    setup_matplotlib_cjk()

    # 构建频率矩阵
    chars = sorted(set("".join(sequences)))
    counts = {c: [0] * n_pos for c in chars}
    for seq in sequences:
        for i, c in enumerate(seq):
            if c in counts:
                counts[c][i] += 1

    counts_df = pd.DataFrame(counts)
    prob_df = counts_df.div(counts_df.sum(axis=1), axis=0)
    info_df = logomaker.transform_matrix(prob_df, from_type="probability", to_type="information")

    # 长序列分块换行：每个分块一行，编号连续
    MAX_BLOCK = 50
    if n_pos <= MAX_BLOCK:
        blocks = [n_pos]
    else:
        blocks = [MAX_BLOCK] * (n_pos // MAX_BLOCK)
        if n_pos % MAX_BLOCK:
            blocks.append(n_pos % MAX_BLOCK)
    n_rows = len(blocks)
    block_w = max(blocks)
    fig, axes = plt.subplots(n_rows, 1, figsize=(max(8, block_w * 0.4), 2.4 * n_rows),
                             squeeze=False)
    axes = axes.flatten()
    ymax = max(float(info_df.max().max()), 4.5)

    for i, blk in enumerate(blocks):
        row_start = sum(blocks[:i])
        sub = info_df.iloc[row_start:row_start + blk]
        logomaker.Logo(sub, ax=axes[i], color_scheme=color_scheme)
        axes[i].set_xlim(row_start - 0.5, row_start + blk - 0.5)
        axes[i].set_xticks(range(row_start, row_start + blk))
        axes[i].set_xticklabels([str(offset + x + 1) for x in range(row_start, row_start + blk)], fontsize=8)
        axes[i].set_ylim(0, ymax)
        axes[i].set_ylabel("bits")
        if n_rows > 1:
            axes[i].set_title(f"位置 {offset + row_start + 1}--{offset + row_start + blk}", fontsize=10, color="#555", pad=8)

    buf = BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight")
    plt.close(fig)
    buf.seek(0)
    img_b64 = base64.b64encode(buf.read()).decode()
    return {
        "image": f"data:image/png;base64,{img_b64}",
        "positions": n_pos,
        "range": [offset + 1, offset + n_pos],
    }


@app.route("/api/weblogo", methods=["POST"])
def api_weblogo():
    """生成序列标识图。结果按请求参数缓存：同一请求并发共享一次渲染，切页回来直接命中缓存秒回。"""
    data = request.get_json()
    sequences = data.get("sequences", [])
    color_scheme = data.get("color_scheme", "chemistry")
    try:
        multimer = int(data.get("multimer") or 1)
    except (TypeError, ValueError):
        multimer = 1
    start = data.get("start")
    end = data.get("end")

    if not sequences or len(sequences) < 2:
        return jsonify({"error": "至少需要 2 条对齐序列"}), 400
    if multimer < 1:
        return jsonify({"error": "多聚体数需 ≥ 1"}), 400

    n_pos = len(sequences[0])
    for s in sequences:
        if len(s) != n_pos:
            return jsonify({"error": "所有序列必须等长（已对齐）"}), 400

    # 多聚体模式：序列为 N 个相同亚基串联，裁剪为单亚基
    if multimer > 1:
        if n_pos % multimer != 0:
            return jsonify({"error": f"序列长度 {n_pos} 不能被多聚体数 {multimer} 整除"}), 400
        n_pos //= multimer
        sequences = [s[:n_pos] for s in sequences]

    # 位点区间（1-based 闭区间）
    offset = 0
    if start is not None or end is not None:
        try:
            lo = int(start) if start is not None else 1
            hi = int(end) if end is not None else n_pos
        except (TypeError, ValueError):
            return jsonify({"error": "位点必须是整数"}), 400
        if lo < 1 or hi > n_pos or lo > hi:
            return jsonify({"error": f"位点区间超出范围（1--{n_pos}）"}), 400
        sequences = [s[lo - 1:hi] for s in sequences]
        n_pos = hi - lo + 1
        offset = lo - 1

    key = hashlib.md5(json.dumps(
        {"seq": sequences, "cs": color_scheme, "mm": multimer, "s": start, "e": end},
        sort_keys=True, ensure_ascii=False).encode("utf-8")).hexdigest()

    with _weblogo_lock:
        if key in _weblogo_cache:
            return jsonify(_weblogo_cache[key])
        ev = _weblogo_inflight.get(key)
        if ev is None:
            ev = _weblogo_inflight[key] = Event()
            owned = True
        else:
            owned = False
    if not owned:
        # 同一请求正在别的线程渲染：等它完成直接取缓存（去重，避免重复计算）
        ev.wait()
        with _weblogo_lock:
            if key in _weblogo_cache:
                return jsonify(_weblogo_cache[key])

    try:
        result = _render_weblogo(sequences, color_scheme, n_pos, offset)
        with _weblogo_lock:
            if len(_weblogo_cache) >= 20:  # 只留最近 20 个结果，防无限增长
                _weblogo_cache.pop(next(iter(_weblogo_cache)))
            _weblogo_cache[key] = result
    finally:
        # 无论成功失败都释放 in-flight，避免后续同请求永久等待
        with _weblogo_lock:
            if key in _weblogo_inflight:
                _weblogo_inflight[key].set()
                del _weblogo_inflight[key]
    return jsonify(result)


# ═══════════════════════════════════════════════════════════
#  Enzyme Activity API
# ═══════════════════════════════════════════════════════════

@app.route("/api/enzyme/parse", methods=["POST"])
def api_enzyme_parse():
    """上传 TECAN xlsx，返回井数据"""
    if "file" not in request.files:
        return jsonify({"error": "请上传 xlsx 文件"}), 400
    f = request.files["file"]
    fd, tmp = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    f.save(tmp)
    try:
        data = parse_tecan_xlsx(tmp)
    except Exception as e:
        return jsonify({"error": f"解析失败: {e}"}), 400
    finally:
        if os.path.exists(tmp):
            os.remove(tmp)
    return jsonify(data)


@app.route("/api/enzyme/fit", methods=["POST"])
def api_enzyme_fit():
    """对指定孔做线性拟合，返回 ΔOD/min + R²；并按阴性/空白孔斜率均值做速率级校正
    （slope_corrected，前端只显示、不再自己算——计算收归后端）。"""
    body = request.get_json()
    wells_data = body.get("wells", {})
    fits = {}
    refs = {}
    for well_id, wd in wells_data.items():
        refs[well_id] = wd.get("ref")
        t = wd.get("times") or []
        o = wd.get("od") or []
        if t and o and len(t) == len(o):
            fits[well_id] = fit_kinetics(t, o)
        else:
            # 空/错位数据：仍返回该孔（null fit）——前端据此覆盖旧 fit，避免筛选成 0 点的孔残留 stale R² 标红
            fits[well_id] = {"slope": None, "intercept": None, "r2": None, "n": len(t)}
    corrected, bg = correct_slopes(fits, refs)
    return jsonify({"wells": corrected, "bg": bg})


@app.route("/api/enzyme/plot", methods=["POST"])
def api_enzyme_plot():
    """生成动力学曲线图，返回 base64 PNG"""
    import base64
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    # 中文字体（内部已 use('Agg') + 注册字体）
    from fonts import setup_matplotlib_cjk
    font_path = setup_matplotlib_cjk()

    body = request.get_json()
    wells_data = body.get("wells", {})
    plot_type = body.get("type", "kinetics")  # kinetics | michaelis
    align_start = body.get("align_start", False)
    align_end = body.get("align_end", False)

    # BLI 绘图风格（配色/字号/脊柱/图例位置）——函数内惰性 import，避免模块顶部拖 scipy
    from bli import COLORS, PLOT_STYLE

    with plt.rc_context(PLOT_STYLE):
        fig, ax = plt.subplots(figsize=(9, 6))

        if plot_type == "kinetics":
            # 扣除阴性/空白 + 对齐起始/终止：纯函数在 calculators 层（sub_blank/align_wells），可单测
            wells_data, mean_neg = sub_blank(wells_data, enabled=body.get("sub_blank", False))
            wells_data, _avg_first, _avg_last = align_wells(wells_data, align_start, align_end)

            # 分组聚合：同组孔逐时间点取平均曲线（aggregate_groups 纯函数），未设组/单孔组照常逐孔画
            error_bar = body.get("error_bar", "sd")  # sd | sem | none
            drawable = {wid: wd for wid, wd in wells_data.items()
                        if wd.get("times") and wd.get("od")
                        and not (wd.get("ref", "") in ("blank", "neg") and not body.get("show_blank", False))}
            groups, singles = aggregate_groups(drawable)

            idx = 0  # 实际绘制条目序号（分组 + 单孔连续计数），用于取 BLI 配色
            plotted_od = []  # 收集实际绘制曲线/拟合线的纵坐标，用于纵轴取整缩放

            # 1) 分组平均曲线 + 误差棒（图例带 (n=成员数)）
            for g in groups:
                label = g["label"]
                conc = g.get("conc_ng_ml")
                lbl = f"{label} (n={g['n']})" + (f" ({conc} ng/mL)" if conc else "")
                times_min = [t / 60 for t in g["times"]]
                od_vals = g["od"]
                color = COLORS[idx % len(COLORS)]
                if error_bar in ("sd", "sem"):
                    yerr = g["err"] if error_bar == "sd" else [e / (g["n"] ** 0.5) for e in g["err"]]
                    ax.errorbar(times_min, od_vals, yerr=yerr, label=lbl, color=color,
                                marker="o", markersize=4, linewidth=1.8, capsize=3, alpha=0.85)
                else:
                    ax.plot(times_min, od_vals, ".-", label=lbl, linewidth=1.8,
                            markersize=4, alpha=0.85, color=color)
                plotted_od.extend(od_vals)
                # 拟合虚线：组内成员平均斜率（slope_corrected 优先），从平均曲线首点出发
                slope = g.get("mean_slope")
                if slope is not None and od_vals:
                    t_fit = np.linspace(times_min[0], times_min[-1], 100)
                    od_fit = [od_vals[0] + slope * (t - t_fit[0]) for t in t_fit]
                    ax.plot(t_fit, od_fit, "--", linewidth=2.3, alpha=0.6, color=color)
                    plotted_od.extend(od_fit)
                idx += 1

            # 2) 单孔曲线（未设组/单孔组，逐孔绘制）
            for well_id in singles:
                wd = drawable[well_id]
                ref = wd.get("ref", "")
                label = wd.get("name", well_id)
                conc = wd.get("conc_ng_ml", "")
                lbl = f"{label}" + (f" ({conc} ng/mL)" if conc else "")
                times_min = [t / 60 for t in wd["times"]]
                od_vals = list(wd["od"])  # 已含 sub_blank + 对齐（见上），直接绘制
                plotted_od.extend(od_vals)

                # 逐孔取 BLI 配色；阳性对照加粗突出
                color = COLORS[idx % len(COLORS)]
                lw = 2.5 if ref == "pos" else 1.8
                alpha_val = 1.0 if ref == "pos" else 0.85
                line = ax.plot(times_min, od_vals, ".-", label=lbl, linewidth=lw,
                               markersize=4, alpha=alpha_val, color=color)
                # 拟合线同色虚线：从曲线首点（od_vals[0]，已含扣减/对齐）出发，斜率优先用扣阴性后的
                # slope_corrected（与表格/已扣曲线一致），无阴性时回退原始 slope
                fit = wd.get("fit")
                slope = (fit.get("slope_corrected") if fit else None)
                if slope is None:
                    slope = fit.get("slope") if fit else None
                if fit and slope is not None:
                    t_fit = np.linspace(times_min[0], times_min[-1], 100)
                    od_fit = [od_vals[0] + slope * (t - t_fit[0]) for t in t_fit]
                    ax.plot(t_fit, od_fit, "--", linewidth=lw + 0.5, alpha=0.6, color=line[0].get_color())
                    plotted_od.extend(od_fit)
                idx += 1

            # 纵轴：数据范围外扩后按数量级取整到整刻度，避免难看的自动刻度
            ylim = snap_ylim(plotted_od)
            if ylim:
                ax.set_ylim(*ylim)

            # 标注：中文/英文自适应
            use_cn = font_path is not None
            notes = []
            if mean_neg is not None:
                notes.append("已扣除阴性/空白" if use_cn else "blank subtracted")
            if align_start:
                notes.append("已对齐起始值" if use_cn else "start aligned")
            if align_end:
                notes.append("已对齐终止值" if use_cn else "end aligned")
            title = "Kinetics"
            if notes:
                title += " (" + ", ".join(notes) + ")"
            ax.set_title(title, fontweight="bold", loc="center")
            ax.set_xlabel("Time (min)")
            ax.set_ylabel("OD")
            if wells_data:
                ax.legend(bbox_to_anchor=(1.02, 1), loc="upper left",
                          frameon=True, fancybox=True, edgecolor="#cccccc")
        elif plot_type == "michaelis":
            points = []
            for well_id, wd in wells_data.items():
                s = wd.get("substrate_uM")
                v = wd.get("rate")
                if s is not None and v is not None:
                    points.append((s, v, wd.get("name", well_id)))
            if points:
                points.sort()
                sx = [p[0] for p in points]
                sy = [p[1] for p in points]
                ax.scatter(sx, sy, c=COLORS[0], s=60, zorder=3)
                for p in points:
                    ax.annotate(p[2], (p[0], p[1]), fontsize=10,
                                textcoords="offset points", xytext=(0, 8))
            ax.set_title("Michaelis-Menten", fontweight="bold", loc="center")
            ax.set_xlabel("Substrate (μM)")
            ax.set_ylabel("Rate (ΔOD/min)")
        ax.grid(alpha=0.15, linestyle="-")
        fig.tight_layout()

    buf = BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight")
    plt.close(fig)
    buf.seek(0)
    return jsonify({"image": f"data:image/png;base64,{base64.b64encode(buf.read()).decode()}"})


@app.route("/api/enzyme/export", methods=["POST"])
def api_enzyme_export():
    """导出酶活原始数据为作图友好 Excel：
    Sheet1「作图数据」宽格式（每孔独立两列：时间 + OD），Sheet2「动力学汇总」（每孔拟合一行）。"""
    body = request.get_json() or {}
    wells = body.get("wells", {}) or {}
    wb = Workbook()

    # Sheet 1: 作图数据（宽格式，每孔时间+OD 两列）
    ws = wb.active
    ws.title = "作图数据"
    headers, col_data = _enzyme_wide_columns(wells)
    _write_wide_ws(ws, headers, col_data)

    # Sheet 2: 动力学汇总
    ws2 = wb.create_sheet("动力学汇总")
    headers2 = ["孔位", "名称", "参考类型", "浓度 (ng/mL)", "浓度 (μM)",
                "斜率 (ΔOD/min)", "截距", "R²", "数据点数"]
    ws2.append(headers2)
    for c in range(1, len(headers2) + 1):
        ws2.cell(row=1, column=c).font = Font(bold=True)
    ref_label = {"blank": "空白", "neg": "阴性", "pos": "阳性"}
    for wid, wd in sorted(wells.items()):
        if not isinstance(wd, dict):
            continue
        fit = wd.get("fit") or {}
        ws2.append([
            wid, wd.get("name", ""), ref_label.get(wd.get("ref", ""), ""),
            wd.get("conc_ng_ml", ""), wd.get("conc_uM", ""),
            fit.get("slope", ""), fit.get("intercept", ""), fit.get("r2", ""),
            len(wd.get("od") or []),
        ])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name="enzyme_well_time_od.xlsx")


# ═══════════════════════════════════════════════════════════
#  BLI 原始数据拟合 API（v0.0.8）
#  ═══════════════════════════════════════════════════════════
# 上传 ForteBio 预处理 CSV → 服务端解析缓存（会话）→ 传感器图 / 5 方法 KD 拟合 /
# 保存为实验（results 带 BLI_ANALYSIS_VERSION；raw→experiment_raw data_type=bli_curves）。
# 曲线数据量较大，不随每次请求回传——解析一次后按 session_id 复用（类 weblogo 缓存）。

_bli_sessions = {}     # session_id -> {"curves": [...], "created": ts}
_bli_lock = Lock()
_BLI_SESSION_TTL = 2 * 3600   # 会话保留 2 小时
_BLI_SESSION_MAX = 10         # 最多保留 10 个会话，防无限增长


def _json_safe(obj):
    """递归把 float 的 NaN/Inf 换成 None——jsonify 序列化非有限数会输出非法 JSON。"""
    import math
    if isinstance(obj, float):
        return None if not math.isfinite(obj) else obj
    if isinstance(obj, np.generic):
        return _json_safe(obj.item())
    if isinstance(obj, dict):
        return {k: _json_safe(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple)):
        return [_json_safe(v) for v in obj]
    return obj


def _bli_curve_to_dict(c) -> dict:
    return {"label": c.label, "sample_id": c.sample_id, "conc_nM": c.conc_nM,
            "time": [float(x) for x in c.time], "response": [float(x) for x in c.response]}


def _bli_dict_to_curve(d: dict):
    from bli import Curve
    return Curve(label=d["label"], sample_id=d["sample_id"], conc_nM=d["conc_nM"],
                 time=np.asarray(d["time"], float), response=np.asarray(d["response"], float))


def _bli_new_session(curves) -> str:
    import uuid
    sid = uuid.uuid4().hex
    now = datetime.now().timestamp()
    with _bli_lock:
        expired = [k for k, v in _bli_sessions.items() if now - v["created"] > _BLI_SESSION_TTL]
        for k in expired:
            del _bli_sessions[k]
        if len(_bli_sessions) >= _BLI_SESSION_MAX:  # 超量丢最旧会话
            oldest = min(_bli_sessions, key=lambda k: _bli_sessions[k]["created"])
            del _bli_sessions[oldest]
        _bli_sessions[sid] = {"curves": [_bli_curve_to_dict(c) for c in curves], "created": now}
    return sid


def _bli_get_session(session_id: str):
    with _bli_lock:
        s = _bli_sessions.get(session_id or "")
        return s["curves"] if s else None


def _bli_filter_curves(curves, active):
    """按曲线 label 过滤（前端勾选「进入数据」的曲线）。active=None → 全保留；
    active=[] → 一条不剩（明确清空）。"""
    if active is None:
        return curves
    active_set = {str(a).strip() for a in active if str(a).strip()}
    return [d for d in curves if d.get("label") in active_set]


def _bli_trim_curves(curves, t_assoc):
    """截去结合起点（t_assoc）之前的数据——默认去掉基线区。返回新 dict 列表。"""
    out = []
    for d in curves:
        t = np.asarray(d["time"], float)
        y = np.asarray(d["response"], float)
        m = t >= t_assoc
        out.append({**d, "time": t[m].tolist(), "response": y[m].tolist()})
    return out


def _bli_opt_float(v):
    if v in (None, ""):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


@app.route("/api/bli/analyze", methods=["POST"])
def api_bli_analyze():
    """上传 ForteBio 预处理 CSV → 解析 → 返回样本摘要 + session_id（后续 plot/fit/save 复用）。"""
    if "file" not in request.files:
        return jsonify({"error": "请上传 ForteBio CSV 文件"}), 400
    f = request.files["file"]
    fd, tmp = tempfile.mkstemp(suffix=".csv")
    os.close(fd)
    f.save(tmp)
    try:
        from bli import parse_fortebio_csv, group_by_sample
        curves = parse_fortebio_csv(tmp)
    except Exception as e:
        return jsonify({"error": f"解析失败: {e}"}), 400
    finally:
        if os.path.exists(tmp):
            os.remove(tmp)
    if not curves:
        return jsonify({"error": "CSV 中无有效曲线数据"}), 400
    sid = _bli_new_session(curves)
    samples = []
    for s_name, s_curves in group_by_sample(curves).items():
        samples.append({
            "sample": s_name,
            "n_curves": len(s_curves),
            "concs": [round(c.conc_nM, 4) for c in s_curves],
            "labels": [c.label for c in s_curves],
        })
    return jsonify({"session_id": sid, "samples": samples, "n_sensors": len(curves)})


@app.route("/api/bli/restore", methods=["POST"])
def api_bli_restore():
    """从实验原始快照重建 BLI 会话（供「从实验复制」把历史实验载回 BLI 分析 tab 再出图/拟合）。

    body: {"payload": {curves: [{label, sample_id, conc_nM, time, response}]}}
    返回形状与 /api/bli/analyze 一致（session_id + samples + n_sensors），前端可直接复用上传路径。
    """
    body = request.get_json() or {}
    curves_raw = (body.get("payload") or {}).get("curves") or []
    if not curves_raw:
        return jsonify({"error": "快照缺少曲线数据"}), 400
    from bli import group_by_sample
    curves = [_bli_dict_to_curve(d) for d in curves_raw]
    sid = _bli_new_session(curves)
    samples = []
    for s_name, s_curves in group_by_sample(curves).items():
        samples.append({
            "sample": s_name,
            "n_curves": len(s_curves),
            "concs": [round(c.conc_nM, 4) for c in s_curves],
            "labels": [c.label for c in s_curves],
        })
    return jsonify({"session_id": sid, "samples": samples, "n_sensors": len(curves)})


@app.route("/api/bli/plot", methods=["POST"])
def api_bli_plot():
    """生成传感器图 PNG（base64）。参数：session_id / smooth_window / fit / t_assoc / t_dissoc /
    separate（每 sample 一图）/ view / mask（sample 过滤）。"""
    body = request.get_json() or {}
    curves = _bli_get_session(body.get("session_id", ""))
    if not curves:
        return jsonify({"error": "会话不存在或已过期，请重新上传"}), 400
    curves = _bli_filter_curves(curves, body.get("active_curves"))
    if not curves:
        return jsonify({"error": "勾选的曲线为空，请重新选择"}), 400
    ta = _bli_opt_float(body.get("t_assoc"))
    td = _bli_opt_float(body.get("t_dissoc"))
    try:
        import base64
        from bli import generate_sensorgram_png, _detect_phases
        # 默认截去结合起点前的基线区（相界在原始曲线上检测，trim 后再出图/叠加拟合）
        if body.get("trim_start", True) is not False:
            if ta is None or td is None:
                a, d = _detect_phases([_bli_dict_to_curve(c) for c in curves])
                ta = ta if ta is not None else a
                td = td if td is not None else d
            curves = _bli_trim_curves(curves, ta)
        curve_objs = [_bli_dict_to_curve(c) for c in curves]
        png = generate_sensorgram_png(
            curve_objs,
            smooth_window=int(body.get("smooth_window", 31) or 0),
            fit=bool(body.get("fit")),
            t_assoc=ta, t_dissoc=td,
            separate=bool(body.get("separate")),
            mask=tuple(body.get("mask") or ()),
            view=tuple(body.get("view") or ()),
        )
        if isinstance(png, dict):
            images = {k: f"data:image/png;base64,{base64.b64encode(v).decode()}"
                      for k, v in png.items()}
            return jsonify({"images": images})
        return jsonify({"image": f"data:image/png;base64,{base64.b64encode(png).decode()}"})
    except Exception as e:
        return jsonify({"error": f"绘图失败: {e}"}), 400


@app.route("/api/bli/fit", methods=["POST"])
def api_bli_fit():
    """对指定 sample 做 5 方法 KD 拟合。参数：session_id / sample / t_assoc / t_dissoc /
    n_concs / no_cutoff / ns_sensor / ns_subtract。"""
    body = request.get_json() or {}
    curves = _bli_get_session(body.get("session_id", ""))
    if not curves:
        return jsonify({"error": "会话不存在或已过期，请重新上传"}), 400
    curves = _bli_filter_curves(curves, body.get("active_curves"))
    sample = body.get("sample", "")
    s_curves = [c for c in (_bli_dict_to_curve(d) for d in curves) if c.sample_id == sample]
    if not s_curves:
        return jsonify({"error": f"找不到样本 {sample}（或该样本曲线已全部取消勾选）"}), 400
    try:
        from bli import fit_kd
        res = fit_kd(
            s_curves,
            t_assoc=_bli_opt_float(body.get("t_assoc")),
            t_dissoc=_bli_opt_float(body.get("t_dissoc")),
            n_concs=int(body.get("n_concs", 8) or 8),
            no_cutoff=bool(body.get("no_cutoff")),
            ns_sensor=body.get("ns_sensor") or None,
            ns_subtract=body.get("ns_subtract", "proportional"),
        )
        return jsonify(_json_safe({"sample": sample, **res}))
    except Exception as e:
        return jsonify({"error": f"拟合失败: {e}"}), 400


@app.route("/api/bli/save", methods=["POST"])
def api_bli_save():
    """保存 BLI 分析为实验：
    - results 带 BLI_ANALYSIS_VERSION（版本契约，可复现规则 #8）
    - 原始曲线落 experiment_raw（data_type=bli_curves，只写一次）
    - 拟合结果由后端按提交参数重算，与用户看到的图/表一致。"""
    body = request.get_json() or {}
    all_curves = _bli_get_session(body.get("session_id", ""))
    if not all_curves:
        return jsonify({"error": "会话不存在或已过期，请重新上传"}), 400
    curves = _bli_filter_curves(all_curves, body.get("active_curves"))
    if not curves:
        return jsonify({"error": "勾选的曲线为空，无法保存"}), 400
    curve_objs = [_bli_dict_to_curve(d) for d in curves]
    try:
        from bli import fit_kd, group_by_sample, BLI_ANALYSIS_VERSION, _detect_phases
    except Exception as e:
        return jsonify({"error": f"BLI 内核加载失败: {e}"}), 500

    # 相界：trim_start（默认开）时若未显式传，用原始曲线自动检测并落绝对时间，
    # 保证存档结果与用户看到的（已截基线）分析一致
    ta = _bli_opt_float(body.get("t_assoc"))
    td = _bli_opt_float(body.get("t_dissoc"))
    trim_start = bool(body.get("trim_start", True))
    if trim_start and (ta is None or td is None):
        a, d = _detect_phases(curve_objs)
        ta = ta if ta is not None else a
        td = td if td is not None else d

    params = {
        "calc_type": "bli_fit",   # 判别字段：详情页渲染 / 从实验复制 / 导出横切依赖（AKTA 同款约定）
        "source": body.get("source", ""),
        "smooth_window": int(body.get("smooth_window", 31) or 0),
        "fit_overlay": bool(body.get("fit_overlay") or body.get("fit")),
        "t_assoc": ta,
        "t_dissoc": td,
        "n_concs": int(body.get("n_concs", 8) or 8),
        "no_cutoff": bool(body.get("no_cutoff")),
        "ns_sensor": body.get("ns_sensor") or None,
        "ns_subtract": body.get("ns_subtract", "proportional"),
        "active_curves": body.get("active_curves") or None,  # 勾选进入数据的曲线 label
        "trim_start": trim_start,                            # 是否截去结合起点前基线
    }

    # 拟合：逐 sample 重算，与前端展示一致
    samples_result = {}
    for s_name, s_curves in group_by_sample(curve_objs).items():
        try:
            r = fit_kd(
                s_curves,
                t_assoc=ta, t_dissoc=td,
                n_concs=params["n_concs"], no_cutoff=params["no_cutoff"],
                ns_sensor=params["ns_sensor"], ns_subtract=params["ns_subtract"],
            )
            samples_result[s_name] = r
        except Exception as e:
            samples_result[s_name] = {"error": str(e)}

    results = {
        "BLI_ANALYSIS_VERSION": BLI_ANALYSIS_VERSION,
        "params": params,
        "samples": samples_result,
    }

    # 原始曲线快照（只写一次，规则 #2/#8）：与实验同事务原子落库（services 统一写入入口），
    # 避免「先建实验再单独 save raw」的部分写入——raw 落库失败不再留孤儿实验。
    # 注意 raw 存**全量**曲线（含未勾选/未截的），active_curves/trim_start 已入 params 可复现。
    # _json_safe 先清 NaN/Inf（float 曲线值），否则 json.dumps 产出非法 JSON 文本。
    raw_payload = _json_safe({
        "analysis_version": BLI_ANALYSIS_VERSION,
        "params": params,
        "curves": list(all_curves),
    })
    try:
        exp = services.create_experiment(
            title=body.get("title", ""),
            exp_type="BLI",
            protein_ids=body.get("protein_ids", []),
            date=body.get("date", ""),
            params=params,
            results=_json_safe(results),
            notes=body.get("notes", ""),
            raw_snapshots=[("bli_curves", raw_payload)],
        )
    except ValueError as err:
        return jsonify({"error": str(err)}), 400
    return jsonify(exp), 201


@app.route("/api/bli/export", methods=["POST"])
def api_bli_export():
    """导出 BLI 分析为 Excel（对标 AKTA 导出）：
    - Sheet1「KD 汇总」：每样品一行 × 5 方法 KD（含备注）
    - Sheet2「作图数据」：每条传感器曲线 = 时间+响应两列，Prism 可直接画传感器图"""
    body = request.get_json() or {}
    curves = _bli_get_session(body.get("session_id", ""))
    if not curves:
        return jsonify({"error": "会话不存在或已过期，请重新上传"}), 400
    curves = _bli_filter_curves(curves, body.get("active_curves"))
    if not curves:
        return jsonify({"error": "勾选的曲线为空，无法导出"}), 400
    try:
        from bli import fit_kd, group_by_sample, _detect_phases
    except Exception as e:
        return jsonify({"error": f"BLI 内核加载失败: {e}"}), 500
    # 默认截去结合起点前基线（相界在原始曲线上检测），作图数据/KD 汇总都基于截后数据
    ta = _bli_opt_float(body.get("t_assoc"))
    td = _bli_opt_float(body.get("t_dissoc"))
    if body.get("trim_start", True) is not False:
        if ta is None or td is None:
            a, d = _detect_phases([_bli_dict_to_curve(c) for c in curves])
            ta = ta if ta is not None else a
            td = td if td is not None else d
        curves = _bli_trim_curves(curves, ta)
    curve_objs = [_bli_dict_to_curve(d) for d in curves]
    fit_params = dict(
        t_assoc=ta, t_dissoc=td,
        n_concs=int(body.get("n_concs", 8) or 8),
        no_cutoff=bool(body.get("no_cutoff")),
        ns_sensor=body.get("ns_sensor") or None,
        ns_subtract=body.get("ns_subtract", "proportional"),
    )

    wb = Workbook()
    # Sheet1 KD 汇总（mixed 无单值 kd，沿用面板约定：稳态优先，否则动力学）
    ws = wb.active
    ws.title = "KD 汇总"
    headers = ["样品", "standard", "split", "joint", "steady", "mixed", "备注"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)
    for s_name, s_curves in group_by_sample(curve_objs).items():
        try:
            res = fit_kd(s_curves, **fit_params)
        except Exception:
            ws.append([s_name, "—", "—", "—", "—", "—", "拟合失败"])
            continue
        row = [s_name]
        notes = []
        for m in ("standard", "split", "joint", "steady", "mixed"):
            v = res.get(m) or {}
            if v.get("error"):
                row.append("—")
                notes.append(f"{m} 拟合失败")
                continue
            if m == "mixed":
                if v.get("kd_steady_mixed") is not None:
                    row.append(round(float(v["kd_steady_mixed"]), 2))
                    notes.append("mixed=稳态")
                elif v.get("kd_kinetic_mixed") is not None:
                    row.append(round(float(v["kd_kinetic_mixed"]), 2))
                    notes.append("mixed=动力学")
                else:
                    row.append("—")
            else:
                kd = v.get("kd")
                row.append(round(float(kd), 2) if kd is not None else "—")
        ws.append(row + ["；".join(notes)])

    # Sheet2 作图数据：每条曲线两列（时间 + 响应），label 撞名时拼 sample_id 去重
    ws2 = wb.create_sheet("作图数据")
    header = []
    seen = set()
    for c in curve_objs:
        lbl = c.label
        if lbl in seen:
            lbl = f"{lbl}·{c.sample_id}"
        seen.add(lbl)
        header += [f"{lbl} 时间 (s)", f"{lbl} 响应 (nm)"]
    _write_wide_ws_pairs(ws2, header, [(c.time, c.response) for c in curve_objs])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True,
                     download_name=f"bli_kd_{body.get('session_id', '')[:8]}.xlsx")


# ═══════════════════════════════════════════════════════════
#  AKTA 峰图整理 API（v0.0.9）
#  ═══════════════════════════════════════════════════════════
# 上传 AKTA Unicorn zip → 标准库原生解析（akta.py，无 pycorn 依赖）→ 通道/事件摘要
# → 峰检测图 / 峰表导出 / 保存为实验（results 带 AKTA_ANALYSIS_VERSION；
#   raw→experiment_raw data_type=akta_traces）。会话缓存类 BLI 分析。

_akta_sessions = {}     # session_id -> {"channels": {...}, "events": {...}, "meta": {...}, "created": ts}
_akta_lock = Lock()
_AKTA_SESSION_TTL = 2 * 3600
_AKTA_SESSION_MAX = 10


def _akta_new_session(parsed: dict) -> str:
    import uuid
    sid = uuid.uuid4().hex
    now = datetime.now().timestamp()
    with _akta_lock:
        expired = [k for k, v in _akta_sessions.items() if now - v["created"] > _AKTA_SESSION_TTL]
        for k in expired:
            del _akta_sessions[k]
        if len(_akta_sessions) >= _AKTA_SESSION_MAX:
            oldest = min(_akta_sessions, key=lambda k: _akta_sessions[k]["created"])
            del _akta_sessions[oldest]
        _akta_sessions[sid] = {**parsed, "created": now}
    return sid


def _akta_get_session(session_id: str):
    with _akta_lock:
        s = _akta_sessions.get(session_id or "")
        return s if s else None


def _akta_auto_title(title: str, source: str) -> str:
    """AKTA 保存默认名：优先 zip 包名（去 .zip 扩展名，保留完整文件名含日期），
    其次系统自动命名（services.create_experiment 的 {date}_AKTA_{seq}）。"""
    title = (title or "").strip()
    if title:
        return title
    src = (source or "").strip()
    if src:
        base = os.path.splitext(os.path.basename(src))[0]
        if base:
            return base
    return ""


@app.route("/api/akta/analyze", methods=["POST"])
def api_akta_analyze():
    """上传 AKTA Unicorn zip（支持多文件批量）→ 逐个解析 →
    返回 {"runs": [{name, session_id, channels, uv_channels, events, meta}]}。"""
    files = request.files.getlist("file") or ([request.files["file"]] if "file" in request.files else [])
    if not files or not files[0].filename:
        return jsonify({"error": "请上传 AKTA Unicorn zip 文件（可多选）"}), 400

    from akta import parse_akta_zip, find_uv_channels
    runs = []
    for f in files:
        if not f.filename:
            continue
        fd, tmp = tempfile.mkstemp(suffix=".zip")
        os.close(fd)
        f.save(tmp)
        try:
            parsed = parse_akta_zip(tmp)
        except Exception as e:
            runs.append({"name": f.filename, "error": f"解析失败: {e}"})
            if os.path.exists(tmp):
                os.remove(tmp)
            continue
        finally:
            if os.path.exists(tmp):
                os.remove(tmp)

        channels = parsed.get("channels", {})
        if not channels:
            runs.append({"name": f.filename, "error": "zip 中未解析出任何通道数据"})
            continue

        sid = _akta_new_session(parsed)
        # 会话里记住源文件名（zip 包名），供绘图标题/图例/导出用
        with _akta_lock:
            _akta_sessions[sid]["source_name"] = f.filename
        channel_list = []
        for name, ch in channels.items():
            channel_list.append({
                "name": name, "data_type": ch.data_type, "unit": ch.unit,
                "n_points": ch.n_points(),
                "vol_start": round(float(ch.vols[0]), 3) if len(ch.vols) else 0,
                "vol_end": round(float(ch.vols[-1]), 3) if len(ch.vols) else 0,
                "amp_min": round(float(np.min(ch.amps)), 3) if len(ch.amps) else 0,
                "amp_max": round(float(np.max(ch.amps)), 3) if len(ch.amps) else 0,
            })
        events_summary = {k: len(v) for k, v in parsed.get("events", {}).items()}
        runs.append({
            "name": f.filename,
            "session_id": sid,
            "channels": channel_list,
            "uv_channels": find_uv_channels(channels),
            "events": events_summary,
            "meta": parsed.get("meta", {}),
        })
    return jsonify({"runs": runs})


@app.route("/api/akta/restore", methods=["POST"])
def api_akta_restore():
    """从实验原始快照重建 AKTA 会话（供「从实验复制」把历史实验载回 AKTA tab 再出图/导出）。

    body: {"payload": {channel:{name,data_type,unit,vols,amps}, events, meta}, "name": 显示名}
    返回形状与 /api/akta/analyze 的单 run 一致（含 session_id），前端可直接塞进 aktaRuns。
    """
    body = request.get_json() or {}
    payload = body.get("payload") or {}
    ch = payload.get("channel") or {}
    vols = ch.get("vols") or []
    if not vols:
        return jsonify({"error": "快照缺少通道曲线数据"}), 400
    from akta import Channel, find_uv_channels
    ch_obj = Channel(name=ch.get("name", "UV"), data_type=ch.get("data_type", "Other"),
                     unit=ch.get("unit", ""), vols=[float(x) for x in vols],
                     amps=[float(x) for x in (ch.get("amps") or [])])
    parsed = {
        "channels": {ch_obj.name: ch_obj},
        "events": payload.get("events") or {},
        "meta": payload.get("meta") or {},
    }
    sid = _akta_new_session(parsed)
    with _akta_lock:
        _akta_sessions[sid]["source_name"] = body.get("name", "") or parsed["meta"].get("run_name", "")
    channel_list = [{
        "name": ch_obj.name, "data_type": ch_obj.data_type, "unit": ch_obj.unit,
        "n_points": ch_obj.n_points(),
        "vol_start": round(float(ch_obj.vols[0]), 3) if len(ch_obj.vols) else 0,
        "vol_end": round(float(ch_obj.vols[-1]), 3) if len(ch_obj.vols) else 0,
        "amp_min": round(float(np.min(ch_obj.amps)), 3) if len(ch_obj.amps) else 0,
        "amp_max": round(float(np.max(ch_obj.amps)), 3) if len(ch_obj.amps) else 0,
    }]
    disp = body.get("name", "") or parsed["meta"].get("run_name", "") or ch_obj.name
    return jsonify({
        "session_id": sid,
        "runs": [{
            "name": disp,
            "session_id": sid,
            "channels": channel_list,
            "uv_channels": find_uv_channels(parsed["channels"]),
            "events": {k: len(v) for k, v in parsed["events"].items()},
            "meta": parsed["meta"],
        }],
    })


@app.route("/api/akta/plot", methods=["POST"])
def api_akta_plot():
    """峰检测 + 峰图 PNG（base64）。参数：session_id / channel / xmin / xmax / min_height /
    smooth_window / show_events（frac 竖线，默认关）/ highlight_frac（目标峰阴影，默认开）/
    target_peak_idx（阴影跟随第几个峰，0=主峰）/ normalize（min-max 归一化，默认关）。
    返回 {image, peaks}。"""
    body = request.get_json() or {}
    sess = _akta_get_session(body.get("session_id", ""))
    if not sess:
        return jsonify({"error": "会话不存在或已过期，请重新上传"}), 400
    ch_name = body.get("channel", "")
    ch = sess["channels"].get(ch_name)
    if ch is None:
        return jsonify({"error": f"找不到通道 {ch_name}"}), 400
    try:
        import base64
        from akta import detect_peaks, generate_akta_png
        xmin = float(body.get("xmin", 0) or 0)
        xmax = body.get("xmax")
        xmax = float(xmax) if xmax not in (None, "") else None
        min_height = float(body.get("min_height", 5) or 5)
        smooth = int(body.get("smooth_window", 11) or 11)
        normalize = bool(body.get("normalize", False))
        events = sess["events"].get("Fraction", [])
        peaks = detect_peaks(ch, xmin=xmin, xmax=xmax, min_height=min_height,
                             smooth_window=smooth)
        try:
            target_peak_idx = int(body.get("target_peak_idx", 0) or 0)
        except (TypeError, ValueError):
            target_peak_idx = 0
        # 图例/标题名：优先前端显式传 sample_name，其次会话里的 zip 包名（去扩展名），再退通道名
        sample_name = (body.get("sample_name") or "").strip()
        if not sample_name:
            src = sess.get("source_name") or ""
            sample_name = os.path.splitext(os.path.basename(src))[0] if src else ""
        png = generate_akta_png(ch, peaks, events=events, xmin=xmin, xmax=xmax,
                                show_events=bool(body.get("show_events", False)),
                                highlight_frac=bool(body.get("highlight_frac", True)),
                                peak_fill=bool(body.get("peak_fill", True)),
                                peak_labels=bool(body.get("peak_labels", False)),
                                target_peak_idx=target_peak_idx,
                                smooth_window=smooth,
                                normalize=normalize,
                                sample_name=sample_name)
        return jsonify({
            "image": f"data:image/png;base64,{base64.b64encode(png).decode()}",
            "peaks": [p.to_dict() for p in peaks],
        })
    except Exception as e:
        return jsonify({"error": f"峰图生成失败: {e}"}), 400


@app.route("/api/akta/overlay", methods=["POST"])
def api_akta_overlay():
    """总图：把多个 session（zip）的指定通道平滑曲线叠在一张图。
    参数：runs: [{session_id, channel, target_peak_idx}], xmin/xmax/smooth_window/
    normalize/show_events/highlight_frac。highlight_frac=True 时每个文件的目标峰
    各自画 frac 阴影（自身+前后各 1 管，颜色跟随曲线）。返回 {image}。"""
    body = request.get_json() or {}
    runs_spec = body.get("runs") or []
    if not runs_spec:
        return jsonify({"error": "请选择至少一个文件"}), 400
    try:
        import base64
        from akta import detect_peaks, generate_akta_overlay_png, \
            fraction_ranges, target_fraction_span
        xmin = float(body.get("xmin", 0) or 0)
        xmax = body.get("xmax")
        xmax = float(xmax) if xmax not in (None, "") else None
        smooth = int(body.get("smooth_window", 11) or 11)
        normalize = bool(body.get("normalize", False))
        show_events = bool(body.get("show_events", False))
        highlight_frac = bool(body.get("highlight_frac", True))
        min_height = float(body.get("min_height", 5) or 5)

        channels, labels, events_union, frac_spans = [], [], [], []
        # 有效 xmax：显式传值或各通道最大体积
        eff_xmax = xmax
        for spec in runs_spec:
            sess = _akta_get_session(spec.get("session_id", ""))
            if not sess:
                continue
            ch = sess["channels"].get(spec.get("channel", ""))
            if ch is None:
                continue
            if len(ch.vols):
                vmax = float(np.max(ch.vols))
                eff_xmax = vmax if eff_xmax is None else max(eff_xmax, vmax)
        for i, spec in enumerate(runs_spec):
            sess = _akta_get_session(spec.get("session_id", ""))
            if not sess:
                continue
            ch = sess["channels"].get(spec.get("channel", ""))
            if ch is None:
                continue
            channels.append(ch)
            src = sess.get("source_name") or ""
            labels.append(os.path.splitext(os.path.basename(src))[0] if src else ch.name)
            frac_events = sess["events"].get("Fraction", [])
            events_union.extend(frac_events)
            # 每个文件的目标峰 frac 阴影：detect 峰 → 目标峰顶点 → 自身+前后 frac
            if highlight_frac:
                try:
                    t_idx = int(spec.get("target_peak_idx", 0) or 0)
                except (TypeError, ValueError):
                    t_idx = 0
                peaks = detect_peaks(ch, xmin=xmin, xmax=xmax, min_height=min_height,
                                     smooth_window=smooth)
                target = peaks[t_idx] if 0 <= t_idx < len(peaks) else (peaks[0] if peaks else None)
                span = (target_fraction_span(fraction_ranges(frac_events, eff_xmax),
                                             target.apex_vol, xmin, eff_xmax)
                        if target else {"self": None})
                frac_spans.append(span)
            else:
                frac_spans.append(None)
        if not channels:
            return jsonify({"error": "所选文件/通道均无效"}), 400
        if len(channels) < 2:
            return jsonify({"error": "总图至少需要 2 个文件的曲线"}), 400

        png = generate_akta_overlay_png(
            channels, events=events_union if show_events else None,
            xmin=xmin, xmax=xmax, show_events=show_events,
            smooth_window=smooth, normalize=normalize, labels=labels,
            frac_spans=frac_spans)
        return jsonify({"image": f"data:image/png;base64,{base64.b64encode(png).decode()}"})
    except Exception as e:
        return jsonify({"error": f"总图生成失败: {e}"}), 400


@app.route("/api/akta/export", methods=["POST"])
def api_akta_export():
    """导出峰表 Excel：
    - Sheet1 峰表（当前 run 的峰）
    - Sheet2 曲线-{channel}（当前 run 单通道 Volume/Signal，保留现状）
    - Sheet3 作图数据（新增）：每个样品（勾选的 run）占两列——体积 + 信号，
      直接可选列作图 / 导入 Prism 对比多条色谱。"""
    body = request.get_json() or {}
    sess = _akta_get_session(body.get("session_id", ""))
    if not sess:
        return jsonify({"error": "会话不存在或已过期，请重新上传"}), 400
    ch_name = body.get("channel", "")
    ch = sess["channels"].get(ch_name)
    if ch is None:
        return jsonify({"error": f"找不到通道 {ch_name}"}), 400
    try:
        from akta import detect_peaks
        xmin = float(body.get("xmin", 0) or 0)
        xmax = body.get("xmax")
        xmax = float(xmax) if xmax not in (None, "") else None
        min_height = float(body.get("min_height", 5) or 5)
        smooth = int(body.get("smooth_window", 11) or 11)
        peaks = detect_peaks(ch, xmin=xmin, xmax=xmax, min_height=min_height,
                             smooth_window=smooth)
    except Exception as e:
        return jsonify({"error": f"峰检测失败: {e}"}), 400

    def _range(v):
        return xmin <= v <= (xmax if xmax is not None else float("inf"))

    wb = Workbook()
    ws = wb.active
    ws.title = "峰表"
    headers = ["峰号", "峰位 (mL)", "峰高 (mAU)", "面积 (mAU·mL)", "起点 (mL)", "终点 (mL)", "半高宽 (mL)"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)
    for i, p in enumerate(peaks, 1):
        ws.append([i, p.apex_vol, p.height, p.area, p.start_vol, p.end_vol, p.half_width])
    ws2 = wb.create_sheet(f"曲线-{ch_name[:20]}")
    ws2.append(["Volume (mL)", "Signal (%s)" % (ch.unit or "")])
    for v, a in zip(ch.vols, ch.amps):
        if _range(v):
            ws2.append([float(v), float(a)])

    # Sheet3 作图数据：每个样品两列（体积 + 信号），跨 run 并列，直接可作图
    runs = body.get("runs") or [{"session_id": body.get("session_id"), "channel": ch_name}]
    samples = []   # [(label, Channel)]
    for rs in runs:
        rsess = _akta_get_session(rs.get("session_id", ""))
        if not rsess:
            continue
        rch = (rsess.get("channels") or {}).get(rs.get("channel", ""))
        if rch is None:
            continue
        label = str(rs.get("name") or rch.name)
        samples.append((label, rch))
    if samples:
        ws3 = wb.create_sheet("作图数据")
        header = []
        col_data = []
        for label, rch in samples:
            header += [f"{label} 体积 (mL)", f"{label} 信号 ({rch.unit or ''})"]
            pairs = [(v, a) for v, a in zip(rch.vols, rch.amps) if _range(v)]
            col_data.append(([p[0] for p in pairs], [p[1] for p in pairs]))
        _write_wide_ws_pairs(ws3, header, col_data)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=f"akta_peaks_{ch_name.replace(' ', '_')}.xlsx")


@app.route("/api/akta/save", methods=["POST"])
def api_akta_save():
    """保存 AKTA 峰图为实验：
    - results 带 AKTA_ANALYSIS_VERSION + 峰表 + 通道摘要
    - 原始通道曲线落 experiment_raw（data_type=akta_traces，只写一次）"""
    body = request.get_json() or {}
    sess = _akta_get_session(body.get("session_id", ""))
    if not sess:
        return jsonify({"error": "会话不存在或已过期，请重新上传"}), 400
    ch_name = body.get("channel", "")
    ch = sess["channels"].get(ch_name)
    if ch is None:
        return jsonify({"error": f"找不到通道 {ch_name}"}), 400
    try:
        from akta import detect_peaks, AKTA_ANALYSIS_VERSION
        xmin = float(body.get("xmin", 0) or 0)
        xmax = body.get("xmax")
        xmax = float(xmax) if xmax not in (None, "") else None
        min_height = float(body.get("min_height", 5) or 5)
        smooth = int(body.get("smooth_window", 11) or 11)
        peaks = detect_peaks(ch, xmin=xmin, xmax=xmax, min_height=min_height,
                             smooth_window=smooth)
    except Exception as e:
        return jsonify({"error": f"峰检测失败: {e}"}), 400

    params = {
        "calc_type": "akta",
        "channel": ch_name,
        "data_type": ch.data_type,
        "unit": ch.unit,
        "xmin": xmin,
        "xmax": xmax,
        "min_height": min_height,
        "smooth_window": smooth,
        "source": body.get("source", ""),
    }
    results = {
        "AKTA_ANALYSIS_VERSION": AKTA_ANALYSIS_VERSION,
        "params": params,
        "n_peaks": len(peaks),
        "peaks": [p.to_dict() for p in peaks],
        "events": {k: len(v) for k, v in sess["events"].items()},
    }

    # 原始曲线快照（只写一次，规则 #2/#8）：与实验同事务原子落库（services 统一写入入口）。
    # _json_safe 先清 NaN/Inf（通道浮点值），否则 json.dumps 产出非法 JSON 文本。
    raw_payload = _json_safe({
        "analysis_version": AKTA_ANALYSIS_VERSION,
        "params": params,
        "channel": ch.to_dict(full=True),
        "events": sess["events"],
        "meta": sess.get("meta", {}),
    })
    try:
        exp = services.create_experiment(
            title=_akta_auto_title(body.get("title", ""), body.get("source", "")),
            exp_type="AKTA",
            protein_ids=body.get("protein_ids", []),
            date=body.get("date", ""),
            params=params,
            results=_json_safe(results),
            notes=body.get("notes", ""),
            raw_snapshots=[("akta_traces", raw_payload)],
        )
    except ValueError as err:
        return jsonify({"error": str(err)}), 400
    return jsonify(exp), 201


def open_browser(port):
    import time
    time.sleep(0.5)  # 等服务器完全就绪
    webbrowser.open(f"http://127.0.0.1:{port}")


def backup_database():
    """启动时自动备份数据库，保留最近 10 份"""
    db_path = models.DB_PATH
    if not os.path.exists(db_path):
        return
    backup_dir = os.path.join(os.path.dirname(db_path), "backups")
    os.makedirs(backup_dir, exist_ok=True)

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = os.path.join(backup_dir, f"protein_lab_{stamp}.db")
    import shutil
    shutil.copy2(db_path, backup_path)

    # 清理旧备份，只保留最近 10 份
    existing = sorted(
        [f for f in os.listdir(backup_dir) if f.endswith(".db")],
        reverse=True,
    )
    for old in existing[10:]:
        os.remove(os.path.join(backup_dir, old))

    print(f"   数据库已备份 -> backups/ ({min(len(existing), 10)} 份)")


if __name__ == "__main__":
    # --mcp: 纯 stdio MCP 模式（须在 print 之前短路，避免污染 JSON-RPC stdout）
    if "--mcp" in sys.argv:
        import mcp_server
        mcp_server.main()
        sys.exit(0)

    # --import-db: 一次性导入旧数据库到 EXE 同目录（仅当目标库无数据时覆盖）
    if "--import-db" in sys.argv:
        try:
            src = sys.argv[sys.argv.index("--import-db") + 1]
        except IndexError:
            src = ""
        if not src or not os.path.exists(src):
            print(f"  源库不存在: {src or '(未提供路径)'}，跳过导入")
        elif bool(models.protein_list() or models.exp_list()):
            print("  目标库已有数据，跳过导入")
        else:
            import shutil
            shutil.copy2(src, models.DB_PATH)
            print(f"  已导入数据库 -> {models.DB_PATH}")

    # 端口：默认 5000，被占用自动顺延找空闲端口；--port <n> 显式指定（占用则报错提示）。
    import socket
    if "--port" in sys.argv:
        try:
            port = int(sys.argv[sys.argv.index("--port") + 1])
        except (IndexError, ValueError):
            print("  --port 需要数字端口，例如 --port 8080")
            sys.exit(1)
        with socket.socket() as s:
            try:
                s.bind(("127.0.0.1", port))
            except OSError:
                print(f"  端口 {port} 已被占用，请用 --port 指定其他端口")
                sys.exit(1)
    else:
        port = None
        for p in range(5000, 5050):
            with socket.socket() as s:
                try:
                    s.bind(("127.0.0.1", p))
                    port = p
                    break
                except OSError:
                    continue
        if port is None:
            print("  找不到可用端口（5000-5049 均被占用）")
            sys.exit(1)

    print("========================================")
    print("  Protein Lab")
    print(f"  服务地址    http://127.0.0.1:{port}")
    print("  浏览器即将自动打开")
    print("  关闭此窗口即停止服务")
    print("========================================")
    backup_database()
    Timer(0.5, open_browser, args=(port,)).start()
    # waitress 生产 WSGI 服务器（纯 Python，跨平台）：无开发服务器警告、不刷请求日志。
    # 静默 waitress 自身的启动/请求日志，只保留我们自己打印的 banner。
    import logging
    logging.getLogger("waitress").setLevel(logging.CRITICAL)
    from waitress import serve
    serve(app, host="127.0.0.1", port=port)
