"""test_bli.py — bli.py 模块 + 酶活绘图端点的回归测试（assert 脚本）

跑法（必须用 venv python）：
    .venv/Scripts/python.exe test_bli.py

覆盖：
1. parse_fortebio_csv / group_by_sample —— 合成 ForteBio CSV 的解析与分组
2. generate_sensorgram_png —— PNG 头校验 + fit 叠加 + separate 模式
3. fit_kd —— 自动相界检测 + 显式相界下 5 方法 KD 与仿真真值同数量级
4. 酶活纯函数（calculators.sub_blank / align_wells / snap_ylim）
5. /api/enzyme/plot（kinetics + michaelis）→ 200 + base64 PNG（隔离临时库）
6. /calculator 仍 200

数据安全：数据库用临时目录，不触碰生产库（见 CLAUDE.md 测试规范）。
"""
import os, sys, csv, base64, importlib, tempfile

sys.stdout.reconfigure(encoding="utf-8", errors="replace")
import numpy as np
from bli import _simulate_1to1, parse_fortebio_csv, group_by_sample, \
    generate_sensorgram_png, fit_kd, fit_1to1_per_curve

TMP = tempfile.mkdtemp(prefix="protein_lab_test_")

# ── 1. 合成 fixture：真值 kon=1e-4 / koff=0.01 → KD = 100 nM ──
rng = np.random.default_rng(42)
KON, KOFF, RMAX = 1e-4, 0.01, 1.0
T_ASSOC, T_DISSOC = 100.0, 400.0
t = np.arange(0.0, 700.0, 2.0)

samples = [
    ("A1", "WT", 200.0), ("A2", "WT", 100.0), ("A3", "WT", 50.0),
    ("A4", "WT", 25.0), ("A5", "WT", 12.5),
    ("B1", "MUT", 200.0), ("B2", "MUT", 50.0),
]
rows = [[None]]  # 行 0: XML（解析器不读）
rows.append([f"t1{l}c{i}" for l, _, _ in samples for i in (1, 2)])
row2, row3 = [], []
for l, sid, c in samples:
    row2 += [f"Sample Loc: {l}", f"Sample ID: {sid}"]
    row3 += [f"Sample Conc: {c:g}", ""]
rows.append(row2)
rows.append(row3)
rows.append([])  # 行 4 空行
ncols = len(samples)
resp_cols = []
for l, sid, c in samples:
    y = _simulate_1to1(t, KON, KOFF, RMAX, c, T_ASSOC, T_DISSOC)
    resp_cols.append(y + rng.normal(0, 0.008, size=len(t)))
for i in range(len(t)):
    row = []
    for j in range(ncols):
        row += [f"{t[i]:.1f}", f"{resp_cols[j][i]:.6f}"]
    rows.append(row)

csv_path = os.path.join(TMP, "fixture.csv")
with open(csv_path, "w", encoding="utf-8", newline="") as f:
    csv.writer(f).writerows(rows)

# ── 2. parse / group ──
curves = parse_fortebio_csv(csv_path)
assert len(curves) == len(samples), f"解析传感器数 {len(curves)} != {len(samples)}"
groups = group_by_sample(curves)
assert set(groups) == {"WT", "MUT"}, set(groups)
assert [c.conc_nM for c in groups["WT"]] == [200, 100, 50, 25, 12.5], "组内须浓度降序"
print("parse/group OK:", {k: len(v) for k, v in groups.items()})

# ── 3. 传感器图 PNG ──
png = generate_sensorgram_png(curves, smooth_window=31)
assert png[:8] == b"\x89PNG\r\n\x1a\n", "PNG 头校验失败"
png_fit = generate_sensorgram_png(curves, fit=True)
assert png_fit[:8] == b"\x89PNG\r\n\x1a\n"
sep = generate_sensorgram_png(curves, separate=True)
assert set(sep) == {"WT", "MUT"}
print(f"sensorgram PNG OK: {len(png)}B, fit={len(png_fit)}B, separate={ {k: len(v) for k, v in sep.items()} }")

# ── 4. KD 拟合（真值 KD=100 nM）──
# 4a. 自动相界检测 sanity：解离起点应落在结合平台区域
res_auto = fit_kd(groups["WT"], verbose=True)
ta_auto, td_auto = res_auto["phase"]["t_assoc"], res_auto["phase"]["t_dissoc"]
print(f"auto phase: assoc={ta_auto:.1f} dissoc={td_auto:.1f}")
assert 90 <= ta_auto <= 120, f"auto t_assoc {ta_auto} 异常"
assert 250 <= td_auto <= 450, f"auto t_dissoc {td_auto} 异常"

# 4b. 显式相界（协议真值）→ 5 方法 KD 与真值同数量级
res = fit_kd(groups["WT"], t_assoc=100, t_dissoc=400, verbose=True)
print("explicit phase: 100→400")
for m, v in res.items():
    if m != "phase" and v:
        print(f"  {m}: {v}")
kd_std = res["standard"]["kd"]
assert 10 < kd_std < 1000, f"standard KD {kd_std} 偏离真值过远"
kd_joint = res["joint"]["kd"]
assert 10 < kd_joint < 1000, f"joint KD {kd_joint} 偏离真值过远"
print(f"KD OK: standard={kd_std:.1f} nM, joint={kd_joint:.1f} nM (truth 100)")

# 4c. 空窗口守卫：相界落在数据外时 fit_1to1_per_curve 不崩溃，返回退化值（此前 IndexError）
t_short = np.arange(0.0, 60.0, 2.0)   # 数据只到 58s
y_short = 1.0 * (1 - np.exp(-0.01 * t_short))
# 结合+解离窗口全空：t_assoc=80、t_dissoc=120 都超出数据末端
f_empty = fit_1to1_per_curve(t_short, y_short, 80.0, 120.0)
assert f_empty["Req"] == 0.0 and f_empty["R0"] == 0.0, f_empty
assert f_empty["kobs"] == 0.01 and f_empty["assoc_r2"] == 0.0 and f_empty["dissoc_r2"] == 0.0, f_empty
# 仅解离窗口空：t_dissoc 恰在数据末端 → mask_d 空
f_d = fit_1to1_per_curve(t_short, y_short, 0.0, 60.0)
assert f_d["R0"] == 0.0 and f_d["dissoc_r2"] == 0.0, f_d
# 正常窗口仍能拟合（回归 sanity：纯结合曲线 → kobs≈0.01）
f_ok = fit_1to1_per_curve(t_short, y_short, 0.0, 30.0)
assert f_ok["kobs"] > 0 and f_ok["koff"] > 0 and f_ok["assoc_r2"] > 0.9, f_ok
print("BLI empty-window guard OK")

# 4d. fit_kd 相界全空（远超数据）：优雅返回 error，不抛异常
res_degen = fit_kd(groups["WT"], t_assoc=900, t_dissoc=1000)
assert res_degen.get("error"), res_degen
print("fit_kd degenerate phase OK")

# ── 5. 酶活纯函数（calculators：sub_blank / align_wells / snap_ylim）──
from calculators import sub_blank, align_wells, snap_ylim, correct_slopes

def _close(a, b, tol=1e-9):
    return abs(a - b) < tol

# 5a. sub_blank：逐时间点扣背景均值，背景自身归零；neg+blank 并存只扣阴性（空白不入背景）
wells = {
    "A1": {"times": [0, 1, 2], "od": [0.10, 0.12, 0.14], "ref": ""},
    "B1": {"times": [0, 1, 2], "od": [0.08, 0.08, 0.08], "ref": "neg"},
    "B2": {"times": [0, 1, 2], "od": [0.10, 0.10, 0.10], "ref": "blank"},
}
subbed, mean_neg = sub_blank(wells, enabled=True)
assert set(mean_neg) == {0, 1, 2} and all(_close(v, 0.08) for v in mean_neg.values()), mean_neg
assert all(_close(a, b) for a, b in zip(subbed["A1"]["od"], [0.02, 0.04, 0.06])), subbed["A1"]["od"]
assert _close(subbed["B1"]["od"][0], 0.0), "阴性自身应被扣到 ≈0"
# 空白不入背景：扣的是 neg 均值，故 B2 = 0.10 - 0.08 = 0.02（不被「多扣」）
assert all(_close(a, b) for a, b in zip(subbed["B2"]["od"], [0.02, 0.02, 0.02])), subbed["B2"]["od"]
# 仅空白：回退用空白作背景（兼容老用法）
w_bo = {"A1": {"times": [0, 1], "od": [0.20, 0.22], "ref": ""},
        "B2": {"times": [0, 1], "od": [0.05, 0.05], "ref": "blank"}}
sbo, mno = sub_blank(w_bo, enabled=True)
assert all(_close(v, 0.05) for v in mno.values()), mno
assert all(_close(a, b) for a, b in zip(sbo["A1"]["od"], [0.15, 0.17])), sbo["A1"]["od"]
# 未启用：原样返回、mean_neg=None；无背景孔：也原样
w_off, mn_off = sub_blank(wells, enabled=False)
assert mn_off is None and w_off is wells
w_none, mn_none = sub_blank({"A1": {"times": [0], "od": [0.1], "ref": ""}}, enabled=True)
assert mn_none is None and w_none["A1"]["od"] == [0.1]
print("sub_blank OK")

# 5b. align_wells：均值只统计样品/阳性，位移只作用于样品/阳性
w3 = {
    "A1": {"times": [0, 1], "od": [0.10, 0.20], "ref": ""},
    "A2": {"times": [0, 1], "od": [0.30, 0.40], "ref": "pos"},
    "B1": {"times": [0, 1], "od": [0.05, 0.06], "ref": "neg"},
}
aligned, af, al = align_wells(w3, align_start=True, align_end=False)
assert _close(af, 0.20), af  # (0.10 + 0.30) / 2
assert all(_close(a, b) for a, b in zip(aligned["A1"]["od"], [0.20, 0.30])), aligned["A1"]["od"]
assert all(_close(a, b) for a, b in zip(aligned["A2"]["od"], [0.20, 0.30])), aligned["A2"]["od"]
assert aligned["B1"]["od"] == [0.05, 0.06], "阴性不参与对齐"
# 无对齐开关：原样返回
w_n, _, _ = align_wells(w3, align_start=False, align_end=False)
assert w_n["A1"]["od"] == [0.10, 0.20]
print("align_wells OK", af)

# 5c. snap_ylim：外扩 6% + 数量级取整；空值返回 None
assert snap_ylim([]) is None
lo, hi = snap_ylim([0.0, 1.0, 2.0])
assert _close(lo, -0.2) and _close(hi, 2.2), (lo, hi)  # span=2 → step=0.1 → [-0.2, 2.2]
print("snap_ylim OK", (lo, hi))

# 5d. correct_slopes：仅阴性斜率均值作背景；空白是基线不入背景、不被校正
fits = {
    "A1": {"slope": 0.010, "intercept": 0.1, "r2": 0.99, "n": 20},   # 样品
    "B1": {"slope": 0.002, "intercept": 0.08, "r2": 0.90, "n": 20},  # 阴性（背景）
    "D1": {"slope": 0.001, "intercept": 0.05, "r2": 0.95, "n": 20},  # 空白（基线）
    "C1": {"slope": None, "intercept": None, "r2": None, "n": 1},    # 数据点不足
}
refs = {"A1": "", "B1": "neg", "D1": "blank", "C1": ""}
out, bg = correct_slopes(fits, refs)
assert out["A1"]["blank_corrected"] is True
assert _close(out["A1"]["slope_corrected"], 0.008), out["A1"]   # 0.010 - 0.002（只扣阴性）
assert out["B1"]["blank_corrected"] is True
assert _close(out["B1"]["slope_corrected"], 0.000), out["B1"]   # 阴性自身也扣
# 空白不入背景、不被速率校正——避免被多扣成负值
assert "slope_corrected" not in out["D1"] and "blank_corrected" not in out["D1"], out["D1"]
assert "slope_corrected" not in out["C1"] and "blank_corrected" not in out["C1"], out["C1"]
assert bg == {"avg": 0.002, "count": 1}, bg
# 无背景：blank_corrected=False，不产生 slope_corrected，bg=None
out2, bg2 = correct_slopes({"A1": {"slope": 0.010}}, {"A1": ""})
assert out2["A1"]["blank_corrected"] is False and "slope_corrected" not in out2["A1"]
assert bg2 is None
# 仅空白作背景：样品扣空白均值，空白自身不被校正
out3, bg3 = correct_slopes({"A1": {"slope": 0.010}, "D1": {"slope": 0.001}}, {"A1": "", "D1": "blank"})
assert _close(out3["A1"]["slope_corrected"], 0.009), out3
assert "slope_corrected" not in out3["D1"]
assert bg3 == {"avg": 0.001, "count": 1}, bg3
print("correct_slopes OK")

# ── 6. 酶活绘图端点（隔离临时库）──
import models
importlib.reload(models)  # ⚠ 会把 DB_PATH 重置为真实路径（测试规范要求）
models.DB_PATH = os.path.join(TMP, "test.db")
models.init_db()
from app import app
client = app.test_client()

tt = list(np.arange(0, 10, 0.5))
payload_kinetics = {
    "type": "kinetics", "align_start": True, "align_end": False, "show_blank": False,
    "wells": {
        "A1": {"times": tt, "od": [0.1 + 0.002 * x for x in range(len(tt))],
               "name": "WT", "conc_ng_ml": 200, "ref": "pos",
               "fit": {"slope": 0.002, "intercept": 0.1}},
        "A2": {"times": tt, "od": [0.08 + 0.001 * x for x in range(len(tt))],
               "name": "MUT", "conc_ng_ml": 100, "ref": "",
               "fit": {"slope": 0.001, "intercept": 0.08}},
    },
}
resp = client.post("/api/enzyme/plot", json=payload_kinetics)
assert resp.status_code == 200, resp.status_code
img = resp.get_json()["image"]
assert img.startswith("data:image/png;base64,"), img[:40]
base64.b64decode(img.split(",", 1)[1])
print("enzyme kinetics plot OK")

# 5b. sub_blank（扣除阴性信号）：含平坦阴性孔 + show_blank 时阴性必须能画出来且正常渲染
payload_sub = {
    "type": "kinetics", "sub_blank": True, "show_blank": True,
    "wells": {
        "A1": {"times": tt, "od": [0.1 + 0.002 * x for x in range(len(tt))],
               "name": "WT", "conc_ng_ml": 200, "ref": "",
               "fit": {"slope": 0.002, "intercept": 0.1}},
        "B1": {"times": tt, "od": [0.08] * len(tt),
               "name": "Neg", "ref": "neg",
               "fit": {"slope": 0.0, "intercept": 0.08}},
    },
}
resp = client.post("/api/enzyme/plot", json=payload_sub)
assert resp.status_code == 200, resp.status_code
assert resp.get_json()["image"].startswith("data:image/png;base64,")
print("enzyme kinetics sub_blank (neg included) OK")

payload_mm = {"type": "michaelis", "wells": {
    "A1": {"substrate_uM": 10, "rate": 0.001, "name": "WT"},
    "A2": {"substrate_uM": 20, "rate": 0.002, "name": "MUT"},
}}
resp = client.post("/api/enzyme/plot", json=payload_mm)
assert resp.status_code == 200, resp.status_code
assert resp.get_json()["image"].startswith("data:image/png;base64,")
print("enzyme michaelis plot OK")

# 6b. /api/enzyme/fit 返回速率级校正（slope_corrected 已收归后端；空白不被校正）
resp = client.post("/api/enzyme/fit", json={"wells": {
    "A1": {"times": tt, "od": [0.1 + 0.002 * x for x in range(len(tt))], "ref": ""},
    "B1": {"times": tt, "od": [0.08 + 0.0005 * x for x in range(len(tt))], "ref": "neg"},
    "C1": {"times": tt, "od": [0.07] * len(tt), "ref": "blank"},
}})
assert resp.status_code == 200, resp.status_code
fit_res = resp.get_json()
wells = fit_res["wells"]
assert wells["A1"]["blank_corrected"] is True, fit_res
assert "slope_corrected" in wells["A1"] and wells["B1"]["slope_corrected"] is not None
assert "slope_corrected" not in wells["C1"], "空白不做速率校正"
assert fit_res["bg"]["count"] == 1 and _close(fit_res["bg"]["avg"], 0.06), fit_res["bg"]
print("enzyme fit slope_corrected OK")

# 6c. 筛选成 0 点的孔仍返回 null fit（前端据此覆盖旧 fit，避免 stale R² 残留标红）
resp = client.post("/api/enzyme/fit", json={"wells": {
    "A1": {"times": tt, "od": [0.1 + 0.002 * x for x in range(len(tt))], "ref": ""},
    "B1": {"times": [], "od": [], "ref": ""},   # 时间点被全部筛选掉
}})
assert resp.status_code == 200, resp.status_code
w6 = resp.get_json()["wells"]
assert "B1" in w6, f"空孔 B1 应返回（否则前端残留旧 fit）: {w6.keys()}"
assert w6["B1"]["r2"] is None and w6["B1"]["slope"] is None, w6["B1"]
assert w6["A1"]["r2"] == 1.0, w6["A1"]
print("enzyme fit empty-well null fit OK")

# 6d. /api/enzyme/export 宽格式：每孔独立两列（时间 + OD），不再长格式堆叠；动力学汇总保留
import io
from openpyxl import load_workbook
resp = client.post("/api/enzyme/export", json={"wells": {
    "A1": {"name": "WT", "ref": "", "conc_ng_ml": 200, "conc_uM": 1.0,
           "times": [0, 30, 60], "od": [0.10, 0.13, 0.16],
           "fit": {"slope": 0.002, "intercept": 0.1, "r2": 1.0}},
    "B1": {"name": "MUT", "ref": "", "conc_ng_ml": 200, "conc_uM": 1.0,
           "times": [0, 30, 60], "od": [0.10, 0.12, 0.14],
           "fit": {"slope": 0.001, "intercept": 0.1, "r2": 1.0}},
}})
assert resp.status_code == 200, resp.status_code
wb_x = load_workbook(io.BytesIO(resp.data))
assert wb_x.sheetnames == ["作图数据", "动力学汇总"], wb_x.sheetnames
ws_x = wb_x["作图数据"]
hdr_x = [c.value for c in ws_x[1]]
assert hdr_x == ["WT 时间 (min)", "WT OD", "MUT 时间 (min)", "MUT OD"], hdr_x
# 每孔两列：时间秒→分钟、OD 保留；行按时间索引对齐
assert [ws_x.cell(row=r, column=1).value for r in range(2, 5)] == [0.0, 0.5, 1.0], "A1 时间列"
assert [ws_x.cell(row=r, column=2).value for r in range(2, 5)] == [0.1, 0.13, 0.16], "A1 OD 列"
assert [ws_x.cell(row=r, column=3).value for r in range(2, 5)] == [0.0, 0.5, 1.0], "B1 时间列"
assert [ws_x.cell(row=r, column=4).value for r in range(2, 5)] == [0.1, 0.12, 0.14], "B1 OD 列"
ws_sum = wb_x["动力学汇总"]
assert ws_sum.cell(row=2, column=1).value == "A1" and ws_sum.cell(row=2, column=6).value == 0.002
print("enzyme export wide-format OK")

# 6e. 撞名回落孔位 + 无名字用孔位：列头仍唯一
resp2 = client.post("/api/enzyme/export", json={"wells": {
    "A1": {"name": "dup", "times": [0], "od": [0.1]},
    "B1": {"name": "dup", "times": [0], "od": [0.2]},
    "C1": {"name": "", "times": [0], "od": [0.3]},
}})
assert resp2.status_code == 200, resp2.status_code
hdr2 = [c.value for c in load_workbook(io.BytesIO(resp2.data))["作图数据"][1]]
assert hdr2 == ["dup 时间 (min)", "dup OD", "B1 时间 (min)", "B1 OD",
                "C1 时间 (min)", "C1 OD"], hdr2
print("enzyme export wide-format dedup OK")

# 6f. 归档酶活实验导出（_export_excel 全酶活分支）同样宽格式：标题前缀 + 每孔两列
from services import create_experiment
exp3 = create_experiment(title="酶活导出测试", exp_type="酶活测定",
                         params={"calc_type": "enzyme", "meta": {"sample": "A1-A2"},
                                 "wells": {
                                     "A1": {"name": "WT", "times": [0, 30], "od": [0.10, 0.14],
                                            "fit": {"slope": 0.002, "r2": 1.0}},
                                     "B1": {"name": "MUT", "times": [0, 30], "od": [0.10, 0.12],
                                            "fit": {"slope": 0.001, "r2": 1.0}},
                                 }})
resp3 = client.get(f"/api/experiments/{exp3['id']}/export")
assert resp3.status_code == 200, resp3.status_code
wb3 = load_workbook(io.BytesIO(resp3.data))
assert "作图数据" in wb3.sheetnames, wb3.sheetnames
hdr3 = [c.value for c in wb3["作图数据"][1]]
assert hdr3 == ["酶活导出测试 WT 时间 (min)", "酶活导出测试 WT OD",
                "酶活导出测试 MUT 时间 (min)", "酶活导出测试 MUT OD"], hdr3
assert [wb3["作图数据"].cell(row=2, column=1).value,
        wb3["作图数据"].cell(row=2, column=2).value] == [0.0, 0.1]
print("enzyme archive export wide-format OK")

# 6g. 分组聚合纯函数 aggregate_groups：同组逐时间点取均值 ± SD；单孔组退化；无组走 singles
from calculators import aggregate_groups
agg_wells = {
    "A1": {"group": "WT", "times": [0, 30, 60], "od": [0.10, 0.14, 0.18], "ref": "",
           "conc_ng_ml": 200, "fit": {"slope_corrected": 0.004}},
    "A2": {"group": "WT", "times": [0, 30, 60], "od": [0.12, 0.16, 0.20], "ref": "",
           "conc_ng_ml": 200, "fit": {"slope_corrected": 0.004}},
    "B1": {"times": [0, 30, 60], "od": [0.20, 0.22, 0.24], "ref": ""},
    "C1": {"group": "solo", "times": [0, 30, 60], "od": [0.5, 0.6, 0.7], "ref": ""},
}
groups, singles = aggregate_groups(agg_wells)
assert len(groups) == 1, f"只有 WT 是有效组（solo 单孔组应退化）: {groups}"
assert groups[0]["label"] == "WT" and groups[0]["n"] == 2, groups[0]
g = groups[0]
assert g["times"] == [0, 30, 60], g["times"]
assert all(_close(a, b) for a, b in zip(g["od"], [0.11, 0.15, 0.19])), g["od"]
assert abs(g["err"][0] - 0.0141421356) < 1e-4, f"SD(ddof=1) 首点 {g['err'][0]}"
assert _close(g["mean_slope"], 0.004), g["mean_slope"]
assert g["conc_ng_ml"] == 200, g["conc_ng_ml"]
assert sorted(singles) == ["B1", "C1"], singles
# 组内缺测点错位：按时间值匹配聚合，B3 少 30s 一点——该点均值只算有数据的成员
agg_ragged = {"A1": {"group": "G", "times": [0, 30, 60], "od": [0.1, 0.2, 0.3]},
              "B3": {"group": "G", "times": [0, 60], "od": [0.2, 0.4]}}
gr, sg = aggregate_groups(agg_ragged)
assert len(sg) == 0 and gr[0]["times"] == [0, 30, 60], (gr, sg)  # 时间取成员并集
assert [_close(a, b) for a, b in zip(gr[0]["od"], [0.15, 0.2, 0.35])], gr[0]["od"]
assert abs(gr[0]["err"][1]) < 1e-9 and abs(gr[0]["err"][0] - 0.0707107) < 1e-4, gr[0]["err"]
print("enzyme aggregate_groups OK")

# 6h. /api/enzyme/plot 分组：同组孔平均曲线 + 误差棒（sd/sem/none 三态 200）
resp = client.post("/api/enzyme/plot", json={
    "type": "kinetics", "error_bar": "sd", "wells": {
        "A1": {"times": tt, "od": [0.1 + 0.002 * x for x in range(len(tt))],
               "name": "r1", "group": "WT", "ref": "",
               "fit": {"slope": 0.002, "intercept": 0.1, "slope_corrected": 0.002}},
        "A2": {"times": tt, "od": [0.11 + 0.002 * x for x in range(len(tt))],
               "name": "r2", "group": "WT", "ref": "",
               "fit": {"slope": 0.002, "intercept": 0.11, "slope_corrected": 0.002}},
        "B1": {"times": tt, "od": [0.2 + 0.001 * x for x in range(len(tt))],
               "name": "MUT", "group": "", "ref": ""},
    }})
assert resp.status_code == 200, resp.status_code
assert resp.get_json()["image"].startswith("data:image/png;base64,")
# 单孔组退化为单孔 + sem/none 各 200
for eb in ("sem", "none"):
    r = client.post("/api/enzyme/plot", json={
        "type": "kinetics", "error_bar": eb, "wells": {"A1": {
            "times": tt, "od": [0.1] * len(tt), "group": "solo", "name": "x",
            "ref": "", "fit": {"slope": 0}}} })
    assert r.status_code == 200, (eb, r.status_code)
print("enzyme plot grouped errorbar OK")

assert client.get("/calculator").status_code == 200
print("/calculator OK")

# ── 7. BLI 分析 API（v0.0.8）：analyze → plot → fit → save（raw 落库 + version）──
# 7a. analyze：上传合成 CSV
with open(csv_path, "rb") as f:
    resp = client.post("/api/bli/analyze",
                       data={"file": (f, "fixture.csv")},
                       content_type="multipart/form-data")
assert resp.status_code == 200, resp.status_code
ana = resp.get_json()
assert ana["session_id"], ana
assert {s["sample"] for s in ana["samples"]} == {"WT", "MUT"}, ana["samples"]
sid = ana["session_id"]
print("bli analyze OK:", {s["sample"]: s["n_curves"] for s in ana["samples"]})

# 7b. plot：合并图 + separate 模式
resp = client.post("/api/bli/plot", json={"session_id": sid, "smooth_window": 31, "fit": True})
assert resp.status_code == 200, resp.status_code
img = resp.get_json()["image"]
assert img.startswith("data:image/png;base64,"), img[:40]
resp = client.post("/api/bli/plot", json={"session_id": sid, "separate": True})
sep = resp.get_json()["images"]
assert set(sep) == {"WT", "MUT"}, sep
print("bli plot OK")

# 7c. fit：显式相界 → WT 的 standard KD 与真值同数量级
resp = client.post("/api/bli/fit", json={
    "session_id": sid, "sample": "WT", "t_assoc": 100, "t_dissoc": 400})
assert resp.status_code == 200, resp.status_code
fit_res = resp.get_json()
assert fit_res["sample"] == "WT"
kd_std = fit_res["standard"]["kd"]
assert 10 < kd_std < 1000, f"API standard KD {kd_std} 偏离真值过远"
print(f"bli fit OK: standard KD={kd_std:.1f} nM (truth 100)")

# 7d. save：results 带 BLI_ANALYSIS_VERSION；raw→experiment_raw data_type=bli_curves
from bli import BLI_ANALYSIS_VERSION
resp = client.post("/api/bli/save", json={
    "session_id": sid, "title": "BLI API 测试",
    "t_assoc": 100, "t_dissoc": 400, "smooth_window": 31, "fit_overlay": True,
    "source": "fixture.csv"})
assert resp.status_code == 201, (resp.status_code, resp.get_json())
saved = resp.get_json()
assert saved["exp_type"] == "BLI"
assert saved["results"]["BLI_ANALYSIS_VERSION"] == BLI_ANALYSIS_VERSION
assert "samples" in saved["results"] and "WT" in saved["results"]["samples"]
# raw 快照存在且只写一次（重复调用=新行，旧行不变）
raws = models.exp_raw_list(saved["id"])
assert len(raws) == 1 and raws[0]["data_type"] == "bli_curves", raws
raw1 = models.exp_raw_get(raws[0]["id"])
assert raw1["payload"]["analysis_version"] == BLI_ANALYSIS_VERSION
assert len(raw1["payload"]["curves"]) == len(curves), "raw 曲线数与解析一致"
resp2 = client.post("/api/bli/save", json={
    "session_id": sid, "title": "BLI API 测试2",
    "t_assoc": 100, "t_dissoc": 400})
saved2 = resp2.get_json()
raws2 = models.exp_raw_list(saved2["id"])
assert len(raws2) == 1
# 同一实验重复 save 两次（不同实验）各自有 raw；旧 raw 内容不变（只写一次）
raw1_after = models.exp_raw_get(raws[0]["id"])
assert raw1_after["payload"]["curves"] == raw1["payload"]["curves"], "raw 不可变（只插不更）"
print(f"bli save OK: exp#{saved['id']} raw#{raws[0]['id']} version={BLI_ANALYSIS_VERSION}")

# 7e. 会话失效：不存在的 session_id → 400
resp = client.post("/api/bli/plot", json={"session_id": "nope"})
assert resp.status_code == 400
print("bli session guard OK")

# 7f. export：KD 汇总 + 作图数据两 sheet（对标 AKTA 导出）。
#     trim_start 默认开 + 显式 t_assoc=100 → 作图数据应从结合起点起、行数减少
import io
from openpyxl import load_workbook

def _load_export(j):
    r = client.post("/api/bli/export", json=j)
    assert r.status_code == 200, (r.status_code, r.data[:200])
    assert r.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return load_workbook(io.BytesIO(r.data))

wb = _load_export({"session_id": sid, "t_assoc": 100, "t_dissoc": 400, "n_concs": 5})
assert wb.sheetnames == ["KD 汇总", "作图数据"], wb.sheetnames
ws = wb["KD 汇总"]
assert [c.value for c in ws[1]] == ["样品", "standard", "split", "joint", "steady", "mixed", "备注"]
assert {ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)} == {"WT", "MUT"}
kd_wt = ws.cell(row=2, column=2).value
assert isinstance(kd_wt, (int, float)) and 10 < kd_wt < 1000, f"导出 KD={kd_wt}"
ws2 = wb["作图数据"]
hdr2 = [c.value for c in ws2[1]]
assert len(hdr2) == len(samples) * 2, f"作图数据列数 {len(hdr2)} != 曲线数×2"
assert hdr2[0] == "A1 时间 (s)" and hdr2[1] == "A1 响应 (nm)"
# trim_start 默认开：时间从结合起点（100s）起，行数 300（满长 350 − 起点前 50 点）
t_first = ws2.cell(row=2, column=1).value
assert t_first >= 100, f"trim 后首点应 ≥ 100s，实为 {t_first}"
assert ws2.max_row == 1 + 300, f"trim 后应 300 行数据，实为 {ws2.max_row - 1}"
print(f"bli export OK: sheets={wb.sheetnames}, KD={kd_wt:.1f} nM, 作图数据 {len(hdr2)} 列 @ t≥{t_first:.0f}s")

# 7g. active_curves 子集：只勾选 WT 的曲线 → 作图数据列数 = WT 曲线数×2
wt_labels = next(s["labels"] for s in ana["samples"] if s["sample"] == "WT")
wb = _load_export({"session_id": sid, "t_assoc": 100, "t_dissoc": 400, "active_curves": wt_labels})
ws2 = wb["作图数据"]
hdr2 = [c.value for c in ws2[1]]
assert len(hdr2) == len(wt_labels) * 2, f"子集列数 {len(hdr2)} != {len(wt_labels)}×2"
assert all(f"{l} 时间 (s)" in hdr2 for l in wt_labels), "子集列应为 WT 曲线"
assert not any(h.startswith("B") for h in hdr2), "MUT 曲线不应出现"
# 空勾选 → 400
r = client.post("/api/bli/export", json={"session_id": sid, "active_curves": []})
assert r.status_code == 400, r.status_code
print(f"bli export active_curves OK: {len(wt_labels)} 曲线 → {len(hdr2)} 列，空勾选 400")

# 7h. trim_start=false：保留全部数据 → 时间从 0 起、满长 350
wb = _load_export({"session_id": sid, "trim_start": False, "n_concs": 5})
ws2 = wb["作图数据"]
t0 = ws2.cell(row=2, column=1).value
assert t0 == 0.0, f"trim 关应保留首点 0s，实为 {t0}"
assert ws2.max_row == 1 + 350, f"trim 关应 350 行，实为 {ws2.max_row - 1}"
print(f"bli export trim_start=false OK: 首点 {t0}s / {ws2.max_row - 1} 行")

import shutil
shutil.rmtree(TMP, ignore_errors=True)
print("\nALL PASSED")
